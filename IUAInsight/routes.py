from datetime import datetime, date, timezone
from flask import render_template, url_for, flash, redirect, request, send_file, abort, after_this_request
from sqlalchemy import and_, or_, case, func, cast, String, extract, text, select
from sqlalchemy.orm import joinedload, selectinload
from flask_login import login_user, current_user, logout_user, login_required
from IUAInsight import app, db, bcrypt
from IUAInsight.services.annee_service import cloture_annee

# ── Modèles OLTP (BD source : lmd1) ────────────────────────────────────────
from IUAInsight.models import (
    Filiere, Specialite, Niveau, Semestre,
    Inscription, Etudiant, AnneeScolaire, Resultat,
    Matiere, Nationalite, Note, Session,
    DetteCreditNiveau, Absence
)

# ── Modèles App (BD applicative : iua_app_db) ──────────────────────────────
from IUAInsight.models_app import (
    Administrateur_sy, Respo_peda,
    Rapport, Alerte 
)

import csv
import zipfile

from IUAInsight.forms import AdminForm, Login_adminForm, UpdateAccountForm, RapportForm, RespoForm
from IUAInsight.ml_models import ml_engine
import os, secrets, statistics, zipfile, subprocess, logging
from PIL import Image
from sqlalchemy import create_engine
from collections import Counter, defaultdict

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from functools import wraps
from io import BytesIO, StringIO

logger = logging.getLogger(__name__)
from sqlalchemy import event

@event.listens_for(Inscription, 'before_insert')
@event.listens_for(Inscription, 'before_update')
def sync_filiere_from_specialite(mapper, connection, target):
    if target.id_specialite and not target.id_filiere:
        sp = Specialite.query.get(target.id_specialite)
        if sp:
            target.id_filiere = sp.id_filiere


# ==========================
# DÉCORATEURS D'ACCÈS
# ==========================
def admin_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not isinstance(current_user, Administrateur_sy):
            abort(403)
        return f(*args, **kwargs)
    return decorated


def respo_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not isinstance(current_user, Respo_peda):
            abort(403)
        return f(*args, **kwargs)
    return decorated


def admin_or_respo_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not isinstance(current_user, (Administrateur_sy, Respo_peda)):
            abort(403)
        return f(*args, **kwargs)
    return decorated


# ==========================
# GESTIONNAIRES D'ERREURS
# ==========================
@app.errorhandler(403)
def forbidden(e):
    return render_template('403.html', title='Accès refusé'), 403

@app.errorhandler(401)
def unauthorized(e):
    return render_template('401.html', title='Non authentifié'), 401


# ==========================
# CONTEXT PROCESSOR
# ==========================
@app.context_processor
@app.context_processor
def inject_globals():
    nb_alertes   = 0
    annee_active = None
    alertes_dropdown = []
    layout       = 'layout.html'

    if current_user.is_authenticated:
        try:
            annee_active = AnneeScolaire.query.filter_by(active=True).first()
        except Exception:
            annee_active = None
        try:
            nb_alertes = db.session.query(func.count(Alerte.id_alerte)).scalar() or 0
            alertes_dropdown = Alerte.query.order_by(Alerte.date.desc()).limit(10).all()
        except Exception:
            nb_alertes = 0

    return dict(nb_alertes=nb_alertes, annee_active=annee_active,
                layout=layout, alertes_dropdown=alertes_dropdown)


# ==========================
# HELPERS
# ==========================
def get_annee_active():
    annee = AnneeScolaire.query.filter_by(active=True).first()
    if not annee:
        annee = AnneeScolaire.query.order_by(AnneeScolaire.id_annee.desc()).first()
    return annee


def get_semestres():
    return [
        {"id_semestre": "1", "libelle": "Semestre 1"},
        {"id_semestre": "2", "libelle": "Semestre 2"},
    ]


def _query_inscriptions_avec_relations(annee_id=None):
    q = Inscription.query.options(
        joinedload(Inscription.etudiant),
        joinedload(Inscription.filiere),
        joinedload(Inscription.specialite),
        joinedload(Inscription.niveau).joinedload(Niveau.niveau_suivant),
        joinedload(Inscription.niveau).selectinload(Niveau.semestres),  # ← FIX statut_simple
        joinedload(Inscription.annee),
        selectinload(Inscription.resultats).joinedload(Resultat.matiere),
        selectinload(Inscription.resultats).joinedload(Resultat.semestre),
    )
    if annee_id:
        q = q.filter(Inscription.id_annee == annee_id)
    return q


def appliquer_filtre_semestre(query, semestre_selected):
    if semestre_selected in ("1", "2"):
        try:
            parite = int(semestre_selected) % 2
            semestres_ids = db.session.query(Semestre.id_semestre)\
                .filter(Semestre.ordre % 2 == parite)\
                .scalar_subquery()
            sous_requete = db.session.query(Resultat.id_inscription)\
                .filter(Resultat.id_semestre.in_(semestres_ids))\
                .scalar_subquery()
            query = query.filter(Inscription.id_inscription.in_(sous_requete))
        except ValueError:
            pass
    return query


def upsert_alerte(type_alerte, msg):
    existante = Alerte.query.filter_by(type_alerte=type_alerte).first()
    if existante:
        existante.message = msg
        existante.date    = datetime.now(timezone.utc)
        existante.vue     = False
    else:
        db.session.add(Alerte(
            type_alerte=type_alerte,
            message=msg,
            date=datetime.now(timezone.utc),
            vue=False,
        ))


def delete_alerte(type_alerte):
    Alerte.query.filter_by(type_alerte=type_alerte).delete()


# ==========================
# MOTEUR ML — HELPERS CENTRAUX
# ==========================
def predict_ml(inscription):
    try:
        return ml_engine.predict(inscription)
    except Exception as e:
        logger.warning("ml_engine.predict() échoué pour inscription %s : %s",
                       inscription.id_inscription, e)
        return None


def predict_batch(inscriptions):
    if not inscriptions:
        return []
    try:
        return ml_engine.predict_batch(inscriptions)
    except Exception as e:
        logger.warning("ml_engine.predict_batch() échoué : %s — fallback individuel", e)
        return [predict_ml(ins) for ins in inscriptions]


def ml_index(inscriptions, ml_results):
    return {
        ins.id_inscription: r
        for ins, r in zip(inscriptions, ml_results)
    }


def ml_badge(ml_result):
    if not ml_result:
        return ("?", "ok")
    niveau = ml_result.get("niveau_risque", "ok")
    MAP = {
        "critique":     ("Critique",     "critique"),
        "modere":       ("Modéré",       "modere"),
        "surveillance": ("Surveillance", "surveillance"),
        "ok":           ("OK",           "ok"),
    }
    return MAP.get(niveau, ("OK", "ok"))


def ml_badge_abandon(ml_result):
    if not ml_result:
        return ("?", "ok")
    niveau = ml_result.get("niveau_abandon", "pas_de_risque")
    MAP = {
        "abandon_probable": ("Abandon probable", "critique"),
        "risque_modere":    ("Risque modéré",    "modere"),
        "pas_de_risque":    ("Pas de risque",    "ok"),
    }
    return MAP.get(niveau, ("Pas de risque", "ok"))


def score_ml_to_int(ml_result):
    if not ml_result:
        return 0
    MAP = {"critique": 80, "modere": 50, "surveillance": 25, "ok": 5}
    return MAP.get(ml_result.get("niveau_risque", "ok"), 0)


# ==========================
# MOTEUR DE SCORE HEURISTIQUE (fallback)
# ==========================
def score_risque(inscription):
    score = 0
    m = inscription.moyenne_annuelle
    if m is None:
        m = inscription.moyenne_s1 or inscription.moyenne_s2

    if m is not None:
        if m < 6:    score += 45
        elif m < 8:  score += 35
        elif m < 10: score += 25
        elif m < 12: score += 10
    else:
        score += 20

    credits_req = inscription.niveau.credits_requis if inscription.niveau else 60
    credits_val = sum(
        r.matiere.credit for r in inscription.resultats
        if r.matiere and r.moyenne is not None and r.moyenne >= 10
    )
    if credits_req > 0:
        ratio = credits_val / credits_req
        if ratio < 0.25:   score += 30
        elif ratio < 0.50: score += 20
        elif ratio < 0.75: score += 10

    if getattr(inscription, 'est_redoublant', False):
        score += 15
    if inscription.moyenne_s1 is None and inscription.moyenne_s2 is None:
        score += 10
    if inscription.moyenne_s1 is not None and inscription.moyenne_s2 is None:
        score += 5

    return min(score, 100)


def niveau_from_score(score):
    if score >= 60: return "critique"
    if score >= 35: return "modere"
    if score >= 15: return "surveillance"
    return "ok"


def badge_risque(score):
    if score >= 60: return ("Critique",     "critique")
    if score >= 35: return ("Modéré",       "modere")
    if score >= 15: return ("Surveillance", "surveillance")
    return              ("OK",           "ok")


# ==========================
# CONSTANTES GENRE
# ==========================
MASCULIN_VALS = {"m", "masculin", "male", "homme", "hommes", "h", "garcon", "garçon"}
FEMININ_VALS  = {"f", "féminin", "feminin", "female", "femme", "femmes", "fille"}

def normaliser_genre(g):
    if not g or g.strip().lower() in ("all", ""):
        return "all"
    g_clean = g.strip().lower()
    if g_clean in MASCULIN_VALS:
        return "M_GROUP"
    if g_clean in FEMININ_VALS:
        return "F_GROUP"
    return "all"


def appliquer_filtre_genre(query, genre_filter):
    if genre_filter == "M_GROUP":
        return query.filter(
            func.lower(func.trim(Etudiant.genre)).in_(list(MASCULIN_VALS))
        )
    if genre_filter == "F_GROUP":
        return query.filter(
            func.lower(func.trim(Etudiant.genre)).in_(list(FEMININ_VALS))
        )
    return query


def _compter_abandon(inscriptions, semestre_filter):
    if semestre_filter == "1":
        return sum(1 for i in inscriptions if i.moyenne_s1 is None)
    if semestre_filter == "2":
        return sum(1 for i in inscriptions if i.moyenne_s2 is None)
    return sum(1 for i in inscriptions if i.moyenne_s1 is None and i.moyenne_s2 is None)


# ===================================================================
# PARTIE RESPONSABLE ACADÉMIQUE
# ===================================================================

# ==========================
# TABLEAU DE BORD
# ==========================
@app.route("/tableau_de_bord")
@respo_required
def tableau_de_bord():
    annee_active    = get_annee_active()
    semestre_filter = request.args.get("semestre", "all")
    genre_filter    = normaliser_genre(request.args.get("genre", "all"))

    base_query = _query_inscriptions_avec_relations(
        annee_id=annee_active.id_annee if annee_active else None
    ).join(Etudiant)
    base_query = appliquer_filtre_genre(base_query, genre_filter)

    inscriptions_toutes = base_query.all()
    inscrits = len(inscriptions_toutes)

    inscriptions = appliquer_filtre_semestre(base_query, semestre_filter).all()

    if semestre_filter == "1":
        moyennes = [i.moyenne_s1 for i in inscriptions if i.moyenne_s1 is not None]
    elif semestre_filter == "2":
        moyennes = [i.moyenne_s2 for i in inscriptions if i.moyenne_s2 is not None]
    else:
        moyennes = [i.moyenne_annuelle for i in inscriptions if i.moyenne_annuelle is not None]

    reussite = sum(1 for m in moyennes if m >= 10)
    echec    = sum(1 for m in moyennes if m <  10)
    abandon  = _compter_abandon(inscriptions_toutes, semestre_filter)

    reussite_pct = round((reussite / inscrits) * 100, 2) if inscrits else 0
    echec_pct    = round((echec    / inscrits) * 100, 2) if inscrits else 0
    abandon_pct  = round((abandon  / inscrits) * 100, 2) if inscrits else 0

    ml_results_promo  = predict_batch(inscriptions)
    ml_niveaux        = [r["niveau_risque"]      for r in ml_results_promo if r]
    ml_probas_abandon = [r["probabilite_abandon"] for r in ml_results_promo if r]
    nb_critiques_ml    = sum(1 for n in ml_niveaux if n == "critique")
    nb_moderes_ml      = sum(1 for n in ml_niveaux if n == "modere")
    nb_surveillance_ml = sum(1 for n in ml_niveaux if n == "surveillance")
    nb_abandons_ml     = sum(1 for p in ml_probas_abandon if p >= 0.7)
    score_abandon_moy  = round(sum(ml_probas_abandon) / len(ml_probas_abandon) * 100, 1) if ml_probas_abandon else 0
    etudiants_risque   = nb_critiques_ml + nb_moderes_ml

    ml_critiques_par_filiere = {}
    for ins, ml_r in zip(inscriptions, ml_results_promo):
        if not ml_r:
            continue
        fid = ins.id_filiere
        if fid not in ml_critiques_par_filiere:
            ml_critiques_par_filiere[fid] = 0
        if ml_r["niveau_risque"] in ("critique", "modere"):
            ml_critiques_par_filiere[fid] += 1

    filiere_agg_q = db.session.query(
        Inscription.id_filiere,
        func.count(Inscription.id_inscription).label("total"),
        func.sum(case((Inscription.moyenne_annuelle >= 10, 1), else_=0)).label("reussis"),
        func.sum(case((Inscription.moyenne_annuelle <  10, 1), else_=0)).label("echecs"),
        func.sum(case((
            Inscription.moyenne_s1.is_(None) if semestre_filter == "1" else (
            Inscription.moyenne_s2.is_(None) if semestre_filter == "2" else
            and_(Inscription.moyenne_s1.is_(None), Inscription.moyenne_s2.is_(None))
            ), 1
        ), else_=0)).label("abandons"),
    ).join(Etudiant, Inscription.id_etudiant == Etudiant.id_etudiant)

    if annee_active:
        filiere_agg_q = filiere_agg_q.filter(Inscription.id_annee == annee_active.id_annee)
    filiere_agg_q = appliquer_filtre_genre(filiere_agg_q, genre_filter)

    if semestre_filter in ("1", "2"):
        try:
            sem = int(semestre_filter)
            sous_requete = db.session.query(Resultat.id_inscription)\
                .filter(Resultat.id_semestre == sem)\
                .scalar_subquery()
            filiere_agg_q = filiere_agg_q.filter(Inscription.id_inscription.in_(sous_requete))
        except ValueError:
            pass

    filiere_agg = filiere_agg_q.group_by(Inscription.id_filiere).all()

    filieres_list = Filiere.query.all()
    filiere_map   = {f.id_filiere: f for f in filieres_list}
    filieres_stats = []
    for row in filiere_agg:
        f = filiere_map.get(row.id_filiere)
        if not f:
            continue
        total_f    = int(row.total    or 0)
        reussite_f = float(row.reussis  or 0)
        echec_f    = float(row.echecs   or 0)
        abandon_f  = float(row.abandons or 0)
        actifs_f   = total_f - abandon_f
        taux_f         = round((reussite_f / actifs_f * 100), 2) if actifs_f else 0
        echec_taux_f   = round((echec_f    / actifs_f * 100), 2) if actifs_f else 0
        abandon_taux_f = round((abandon_f  / total_f  * 100), 2) if total_f  else 0
        nb_ml_risque_f = ml_critiques_par_filiere.get(row.id_filiere, 0)
        filieres_stats.append({
            "nom":          f.nom_filiere,
            "etudiants":    total_f,
            "reussite":     taux_f,
            "echec":        echec_taux_f,
            "abandon":      abandon_taux_f,
            "nb_ml_risque": nb_ml_risque_f,
            "statut": (
                "Excellent" if taux_f >= 80 else
                "Bon"       if taux_f >= 60 else
                "Moyen"     if taux_f >= 40 else
                "Critique"
            )
        })

    annees_triees    = AnneeScolaire.query.order_by(AnneeScolaire.date_debut.desc()).all()
    annee_precedente = None
    if annee_active:
        for idx, a in enumerate(annees_triees):
            if a.id_annee == annee_active.id_annee and idx + 1 < len(annees_triees):
                annee_precedente = annees_triees[idx + 1]
                break

    def _to_float(v):
        return float(v) if v is not None else 0.0

    def calc_stats_annee(id_annee):
        base = db.session.query(
            func.count(Inscription.id_inscription).label("total"),
        ).join(Etudiant, Inscription.id_etudiant == Etudiant.id_etudiant)\
         .filter(Inscription.id_annee == id_annee)
        base = appliquer_filtre_genre(base, genre_filter)
        total_p = int(base.scalar() or 0)

        if semestre_filter == "1":
            moy_col = Inscription.moyenne_s1
        elif semestre_filter == "2":
            moy_col = Inscription.moyenne_s2
        else:
            moy_col = Inscription.moyenne_annuelle

        agg = db.session.query(
            func.sum(case((moy_col >= 10, 1), else_=0)).label("reussis"),
            func.sum(case((moy_col <  10, 1), else_=0)).label("echecs"),
            func.sum(case((
                and_(Inscription.moyenne_s1.is_(None), Inscription.moyenne_s2.is_(None))
                if semestre_filter == "all" else
                (Inscription.moyenne_s1.is_(None) if semestre_filter == "1" else Inscription.moyenne_s2.is_(None))
            , 1), else_=0)).label("abandons"),
        ).join(Etudiant, Inscription.id_etudiant == Etudiant.id_etudiant)\
         .filter(Inscription.id_annee == id_annee, moy_col.isnot(None))
        agg = appliquer_filtre_genre(agg, genre_filter)
        row = agg.first()

        reussis_p = _to_float(row.reussis  if row else 0)
        echecs_p  = _to_float(row.echecs   if row else 0)
        abandons_p= _to_float(row.abandons if row else 0)
        actifs_p  = reussis_p + echecs_p
        reussite_p = round(reussis_p  / actifs_p * 100, 2) if actifs_p else 0.0
        echec_p    = round(echecs_p   / actifs_p * 100, 2) if actifs_p else 0.0
        abandon_p  = round(abandons_p / total_p  * 100, 2) if total_p  else 0.0
        return {
            "inscrits": total_p,
            "reussite": reussite_p,
            "echec":    echec_p,
            "abandon":  abandon_p,
        }

    stats_prec = calc_stats_annee(annee_precedente.id_annee) if annee_precedente else None

    def calc_delta(val_actuelle, val_prec):
        if val_prec is None:
            return None, "stable"
        d = round(float(val_actuelle) - float(val_prec), 1)
        return d, ("up" if d > 0 else "down" if d < 0 else "stable")

    delta_inscrits, sens_inscrits = calc_delta(inscrits,     stats_prec["inscrits"] if stats_prec else None)
    delta_reussite, sens_reussite = calc_delta(reussite_pct, stats_prec["reussite"] if stats_prec else None)
    delta_echec,    sens_echec    = calc_delta(echec_pct,    stats_prec["echec"]    if stats_prec else None)
    delta_abandon,  sens_abandon  = calc_delta(abandon_pct,  stats_prec["abandon"]  if stats_prec else None)

    meilleure_filiere  = max(filieres_stats, key=lambda x: x["reussite"])["nom"] if filieres_stats else "-"
    nb_filieres        = len(filieres_stats)
    critiques_filieres = sum(1 for row in filiere_agg if (row.total or 0) > 0 and (float(row.echecs or 0) / float(row.total)) >= 0.4)
    moderees_filieres  = sum(1 for row in filiere_agg if (row.total or 0) > 0 and 0.3 <= (float(row.echecs or 0) / float(row.total)) < 0.4)

    niveau_agg_tb = db.session.query(
        Inscription.id_niveau,
        func.count(Inscription.id_inscription).label("total"),
        func.sum(case((Inscription.moyenne_annuelle < 10, 1), else_=0)).label("echecs"),
    ).join(Etudiant, Inscription.id_etudiant == Etudiant.id_etudiant)
    if annee_active:
        niveau_agg_tb = niveau_agg_tb.filter(Inscription.id_annee == annee_active.id_annee)
    niveau_agg_tb     = appliquer_filtre_genre(niveau_agg_tb, genre_filter)
    niveau_agg_tb     = niveau_agg_tb.group_by(Inscription.id_niveau).all()
    niveaux_critiques = sum(1 for row in niveau_agg_tb if (row.total or 0) > 0 and (float(row.echecs or 0) / float(row.total)) >= 0.4)

    if echec_pct > 30:
        upsert_alerte("echec_global_eleve", f"Taux d'échec global élevé : {echec_pct}%")
    else:
        delete_alerte("echec_global_eleve")
    if abandon_pct > 15:
        upsert_alerte("abandon_en_hausse", f"Abandon en hausse : {abandon_pct}%")
    else:
        delete_alerte("abandon_en_hausse")
    if nb_critiques_ml > 0:
        upsert_alerte("ml_critiques_promo",
            f"ML détecte {nb_critiques_ml} étudiant(s) en risque critique · {nb_abandons_ml} à fort risque d'abandon")
    else:
        delete_alerte("ml_critiques_promo")

    db.session.commit()

    now_str  = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    alertes  = [
        {"message": a.message, "date": a.date.strftime("%Y-%m-%d") if a.date else now_str}
        for a in Alerte.query.order_by(Alerte.date.desc()).limit(10).all()
    ]
    nb_alertes = db.session.query(func.count(Alerte.id_alerte)).scalar() or 0
    semestres  = get_semestres()

    genre_filter_label = (
        "Masculin" if genre_filter == "M_GROUP" else
        "Féminin"  if genre_filter == "F_GROUP" else
        "all"
    )

    return render_template(
        "respo/tableau_de_bord.html",
        reussite=reussite_pct, echec=echec_pct, abandon=abandon_pct,
        inscrits=inscrits, etudiants_risque=etudiants_risque,
        meilleure_filiere=meilleure_filiere, filieres_stats=filieres_stats,
        alertes=alertes, nb_filieres=nb_filieres, nb_alertes=nb_alertes,
        critiques=critiques_filieres, moderees=moderees_filieres,
        niveaux_critiques=niveaux_critiques, semestre=semestres,
        semestre_selected=semestre_filter, genre_filter=genre_filter_label,
        annee_active=annee_active, nb_critiques_ml=nb_critiques_ml,
        nb_moderes_ml=nb_moderes_ml, nb_surveillance_ml=nb_surveillance_ml,
        nb_abandons_ml=nb_abandons_ml, score_abandon_moy=score_abandon_moy,
        ml_source=ml_engine.status().get("risk_model_trained") and "ml" or "heuristic",
        annee_precedente=annee_precedente,
        delta_inscrits=delta_inscrits,  sens_inscrits=sens_inscrits,
        delta_reussite=delta_reussite,  sens_reussite=sens_reussite,
        delta_echec=delta_echec,        sens_echec=sens_echec,
        delta_abandon=delta_abandon,    sens_abandon=sens_abandon,
        title='Tableau de bord global'
    )


# ==========================
# TABLEAU SPÉCIALITÉ
# ==========================
@app.route('/tableau_sp', methods=["GET"])
@respo_required
def tableau_sp():
    annee_active = get_annee_active()

    filiere_selected  = request.args.get("filiere",  "all")
    niveau_selected   = request.args.get("niveau",   "all")
    semestre_selected = request.args.get("semestre", "all")

    filieres    = Filiere.query.order_by(Filiere.nom_filiere).all()
    specialites = Specialite.query.order_by(Specialite.nom_specialite).all()
    niveaux     = Niveau.query.order_by(Niveau.libelle).all()
    semestres   = get_semestres()

    query = _query_inscriptions_avec_relations(
        annee_id=annee_active.id_annee if annee_active else None
    )
    try:
        if filiere_selected != "all":
            query = query.filter(Inscription.id_filiere == int(filiere_selected))
        if niveau_selected != "all":
            query = query.filter(Inscription.id_niveau == int(niveau_selected))
    except ValueError:
        pass
    inscriptions = query.all()

    ml_results         = predict_batch(inscriptions)
    ml_par_inscription = ml_index(inscriptions, ml_results)

    groupes = []
    if filiere_selected != "all":
        try:
            fid = int(filiere_selected)
            sps = [s for s in specialites if s.id_filiere == fid]
            filiere_obj = next((f for f in filieres if f.id_filiere == fid), None)
            if sps:
                for sp in sps:
                    groupes.append({"label": sp.nom_specialite, "id_specialite": sp.id_specialite, "id_filiere": None})
                sans_sp = [i for i in inscriptions if i.id_filiere == fid and i.id_specialite is None]
                if sans_sp:
                    groupes.append({
                        "label":         f"{filiere_obj.nom_filiere} (sans spécialité)" if filiere_obj else "Sans spécialité",
                        "id_specialite": None,
                        "id_filiere":    fid,
                    })
            else:
                groupes.append({
                    "label":         filiere_obj.nom_filiere if filiere_obj else "Sans spécialité",
                    "id_specialite": None,
                    "id_filiere":    fid,
                })
        except ValueError:
            pass
    else:
        for sp in specialites:
            groupes.append({"label": sp.nom_specialite, "id_specialite": sp.id_specialite, "id_filiere": None})
        ids_filiere_avec_sp = {s.id_filiere for s in specialites}
        for f in filieres:
            if f.id_filiere in ids_filiere_avec_sp:
                sans_sp = [i for i in inscriptions if i.id_filiere == f.id_filiere and i.id_specialite is None]
                if sans_sp:
                    groupes.append({"label": f"{f.nom_filiere} (sans spécialité)", "id_specialite": None, "id_filiere": f.id_filiere})
            else:
                groupes.append({"label": f.nom_filiere, "id_specialite": None, "id_filiere": f.id_filiere})

    liste = []
    for groupe in groupes:
        if groupe["id_specialite"] is not None:
            insc_grp = [i for i in inscriptions if i.id_specialite == groupe["id_specialite"]]
        elif groupe["id_filiere"] is not None:
            insc_grp = [i for i in inscriptions if i.id_filiere == groupe["id_filiere"] and i.id_specialite is None]
        else:
            insc_grp = [i for i in inscriptions if i.id_specialite is None]

        niveaux_a_afficher = niveaux
        if niveau_selected != "all":
            try:
                niveaux_a_afficher = [n for n in niveaux if n.id_niveau == int(niveau_selected)]
            except ValueError:
                pass

        stats_niveaux = []
        for niv in niveaux_a_afficher:
            insc_niv = [i for i in insc_grp if i.id_niveau == niv.id_niveau]
            inscrits = len(insc_niv)

            if semestre_selected == "1":
                admis        = sum(1 for i in insc_niv if i.moyenne_s1 is not None and i._credits_s1_apres_rattrapage() >= i._credits_par_semestre())
                admis_dettes = None
                redoublants  = None
                ajournes     = sum(1 for i in insc_niv if i.moyenne_s1 is not None and i._credits_s1_apres_rattrapage() < i._credits_par_semestre())

            elif semestre_selected == "2":
                admis        = sum(1 for i in insc_niv if i.moyenne_s2 is not None and i._credits_s2_apres_rattrapage() >= i._credits_par_semestre())
                admis_dettes = None
                redoublants  = None
                ajournes     = sum(1 for i in insc_niv if i.moyenne_s2 is not None and i._credits_s2_apres_rattrapage() < i._credits_par_semestre())

            else:
                redoublants  = sum(1 for i in insc_niv if i.statut_simple == "Redoublant")
                admis        = sum(1 for i in insc_niv if i.statut_simple == "Admis")
                admis_dettes = sum(1 for i in insc_niv if i.statut_simple == "Admis (dettes)")
                ajournes     = sum(1 for i in insc_niv if "Ajourné" in i.statut_simple)

            presents = admis + (admis_dettes or 0) + ajournes + (redoublants or 0)

            ml_critiques_niv = sum(
                1 for i in insc_niv
                if ml_par_inscription.get(i.id_inscription)
                and ml_par_inscription[i.id_inscription]["niveau_risque"] in ("critique", "modere")
            )
            ml_abandons_niv = sum(
                1 for i in insc_niv
                if ml_par_inscription.get(i.id_inscription)
                and ml_par_inscription[i.id_inscription]["probabilite_abandon"] >= 0.7
            )

            stats_niveaux.append({
                "niveau":        niv.libelle,
                "inscrits":      inscrits,
                "admis":         admis,
                "admis_dettes":  admis_dettes,
                "ajournes":      ajournes,
                "redoublants":   redoublants,
                "taux_reussite": round((admis + (admis_dettes or 0)) / presents * 100, 2) if presents else 0,
                "taux_echec":    round((ajournes + (redoublants or 0)) / presents * 100, 2) if presents else 0,
                "ml_critiques":  ml_critiques_niv,
                "ml_abandons":   ml_abandons_niv,
            })

        liste.append({"specialite": groupe["label"], "stats_niveaux": stats_niveaux})

    liste_flat = []
    for sp in liste:
        for niv in sp["stats_niveaux"]:
            if niv["inscrits"] > 0:
                liste_flat.append({
                    "specialite":    sp["specialite"],
                    "niveau":        niv["niveau"],
                    "inscrits":      niv["inscrits"],
                    "admis":         niv["admis"],
                    "admis_dettes":  niv["admis_dettes"],
                    "ajournes":      niv["ajournes"],
                    "redoublants":   niv["redoublants"],
                    "taux_reussite": niv["taux_reussite"],
                    "taux_echec":    niv["taux_echec"],
                    "ml_critiques":  niv["ml_critiques"],
                    "ml_abandons":   niv["ml_abandons"],
                })

    liste_flat.sort(key=lambda x: (x["taux_reussite"], -x["taux_echec"], x["inscrits"]), reverse=True)

    return render_template(
        "respo/tableau_sp.html",
        filieres=filieres, specialites=specialites, niveaux=niveaux,
        semestre=semestres, filiere_selected=filiere_selected,
        niveau_selected=niveau_selected, semestre_selected=semestre_selected,
        liste=liste_flat, annee_active=annee_active,
        title="Comparaison par spécialité et niveau"
    )
# ==========================
# TABLEAU FILIÈRE
# ==========================
@app.route("/tableau_f", methods=["GET", "POST"])
@respo_required
def tableau_f():
    annee_active = get_annee_active()

    semestre_selected = request.args.get("semestre", "all")
    filiere_selected  = request.args.get("filiere",  "all")
    niveau_selected   = request.args.get("niveau",   "all")

    filieres  = Filiere.query.order_by(Filiere.nom_filiere).all()
    niveaux   = Niveau.query.order_by(Niveau.libelle).all()
    semestres = get_semestres()

    query_total = _query_inscriptions_avec_relations(
        annee_id=annee_active.id_annee if annee_active else None
    ).join(Filiere).join(Niveau)
    try:
        if filiere_selected and filiere_selected != "all":
            query_total = query_total.filter(Inscription.id_filiere == int(filiere_selected))
        if niveau_selected and niveau_selected != "all":
            query_total = query_total.filter(Inscription.id_niveau == int(niveau_selected))
    except ValueError:
        pass

    ins_all      = query_total.all()
    total        = len(ins_all)
    query        = appliquer_filtre_semestre(query_total, semestre_selected)
    inscriptions = query.all()
    total_actifs = len(inscriptions)

    if semestre_selected == "1":
        reussis = sum(1 for i in inscriptions if i.moyenne_s1 and i.moyenne_s1 >= 10)
        echecs  = sum(1 for i in inscriptions if i.moyenne_s1 and i.moyenne_s1 < 10)
    elif semestre_selected == "2":
        reussis = sum(1 for i in inscriptions if i.moyenne_s2 and i.moyenne_s2 >= 10)
        echecs  = sum(1 for i in inscriptions if i.moyenne_s2 and i.moyenne_s2 < 10)
    else:
        reussis = sum(1 for i in inscriptions if i.moyenne_annuelle and i.moyenne_annuelle >= 10)
        echecs  = sum(1 for i in inscriptions if i.moyenne_annuelle and i.moyenne_annuelle < 10)

    taux_reussite    = round((reussis / total_actifs * 100), 2) if total_actifs else 0
    taux_echec       = round((echecs  / total_actifs * 100), 2) if total_actifs else 0
    moyenne_generale = round(
        sum(i.moyenne_annuelle for i in inscriptions if i.moyenne_annuelle is not None) /
        max(1, sum(1 for i in inscriptions if i.moyenne_annuelle is not None)), 2
    ) if total_actifs else 0

    sans_moyenne = sum(1 for i in ins_all if i.moyenne_s1 is None and i.moyenne_s2 is None)
    taux_e       = round((sans_moyenne / total * 100), 2) if total else 0

    ml_results        = predict_batch(inscriptions)
    ml_niveaux        = [r["niveau_risque"]      for r in ml_results if r]
    ml_probas         = [r["probabilite_abandon"] for r in ml_results if r]
    nb_critiques_ml   = sum(1 for n in ml_niveaux if n == "critique")
    nb_moderes_ml     = sum(1 for n in ml_niveaux if n == "modere")
    nb_abandons_ml    = sum(1 for p in ml_probas  if p >= 0.7)
    proba_abandon_moy = round(sum(ml_probas) / len(ml_probas) * 100, 1) if ml_probas else 0

    agg_niveaux_q = db.session.query(
        Inscription.id_niveau,
        func.count(Inscription.id_inscription).label("total"),
        func.sum(case(
            (Inscription.moyenne_s1 >= 10, 1) if semestre_selected == "1" else
            (Inscription.moyenne_s2 >= 10, 1) if semestre_selected == "2" else
            (Inscription.moyenne_annuelle >= 10, 1),
            else_=0
        )).label("reussis"),
        func.sum(case(
            (Inscription.moyenne_s1 < 10, 1) if semestre_selected == "1" else
            (Inscription.moyenne_s2 < 10, 1) if semestre_selected == "2" else
            (Inscription.moyenne_annuelle < 10, 1),
            else_=0
        )).label("echecs"),
    )
    if annee_active:
        agg_niveaux_q = agg_niveaux_q.filter(Inscription.id_annee == annee_active.id_annee)
    try:
        if filiere_selected != "all":
            agg_niveaux_q = agg_niveaux_q.filter(Inscription.id_filiere == int(filiere_selected))
        if niveau_selected != "all":
            agg_niveaux_q = agg_niveaux_q.filter(Inscription.id_niveau == int(niveau_selected))
    except ValueError:
        pass
    agg_niveaux = agg_niveaux_q.group_by(Inscription.id_niveau).all()

    niveaux_map         = {n.id_niveau: n.libelle for n in niveaux}
    reussite_par_niveau = {}
    echec_par_niveau    = {}
    for row in agg_niveaux:
        libelle  = niveaux_map.get(row.id_niveau, str(row.id_niveau))
        total_n  = int(row.total   or 0)
        reussi_n = float(row.reussis or 0)
        echec_n  = float(row.echecs  or 0)
        reussite_par_niveau[libelle] = round((reussi_n / total_n * 100), 2) if total_n else 0
        echec_par_niveau[libelle]    = round((echec_n  / total_n * 100), 2) if total_n else 0

    annees = AnneeScolaire.query.order_by(AnneeScolaire.libelle).limit(5).all()

    filieres_graphique = filieres
    if filiere_selected != "all":
        try:
            fid = int(filiere_selected)
            filieres_graphique = [f for f in filieres if f.id_filiere == fid]
        except ValueError:
            pass

    annee_ids_graph    = [a.id_annee for a in annees]
    filieres_ids_graph = [f.id_filiere for f in filieres_graphique]

    agg_evo_q = db.session.query(
        Inscription.id_filiere,
        Inscription.id_annee,
        func.count(Inscription.id_inscription).label("total"),
        func.sum(case((Inscription.moyenne_annuelle >= 10, 1), else_=0)).label("reussi"),
    ).filter(
        Inscription.id_filiere.in_(filieres_ids_graph),
        Inscription.id_annee.in_(annee_ids_graph),
    )
    if niveau_selected != "all":
        try:
            agg_evo_q = agg_evo_q.filter(Inscription.id_niveau == int(niveau_selected))
        except ValueError:
            pass
    agg_evo = agg_evo_q.group_by(Inscription.id_filiere, Inscription.id_annee).all()
    evo_index = {(r.id_filiere, r.id_annee): r for r in agg_evo}

    evolution_par_filiere = {}
    for f in filieres_graphique:
        valeurs = []
        for a in annees:
            row     = evo_index.get((f.id_filiere, a.id_annee))
            total_a = int(row.total  if row else 0)
            reussi  = float(row.reussi if row else 0)
            valeurs.append(round(reussi / total_a * 100, 2) if total_a else 0)
        evolution_par_filiere[f.nom_filiere] = valeurs

    annees_triees    = AnneeScolaire.query.order_by(AnneeScolaire.date_debut.desc()).all()
    annee_precedente = None
    if annee_active:
        for idx, a in enumerate(annees_triees):
            if a.id_annee == annee_active.id_annee and idx + 1 < len(annees_triees):
                annee_precedente = annees_triees[idx + 1]
                break

    def calc_stats_annee_f(id_annee):
        filters = [Inscription.id_annee == id_annee]
        try:
            if filiere_selected != "all":
                filters.append(Inscription.id_filiere == int(filiere_selected))
            if niveau_selected != "all":
                filters.append(Inscription.id_niveau == int(niveau_selected))
        except ValueError:
            pass

        total_p = int(db.session.query(func.count(Inscription.id_inscription))
                      .filter(*filters).scalar() or 0)

        if semestre_selected == "1":
            moy_col = Inscription.moyenne_s1
        elif semestre_selected == "2":
            moy_col = Inscription.moyenne_s2
        else:
            moy_col = Inscription.moyenne_annuelle

        agg = db.session.query(
            func.sum(case((moy_col >= 10, 1), else_=0)).label("reussis"),
            func.sum(case((moy_col <  10, 1), else_=0)).label("echecs"),
            func.avg(moy_col).label("moy"),
            func.sum(case((and_(Inscription.moyenne_s1.is_(None), Inscription.moyenne_s2.is_(None)), 1), else_=0)).label("abandons"),
        ).filter(*filters, moy_col.isnot(None)).first()

        reussis_f  = float(agg.reussis  or 0) if agg else 0.0
        echecs_f   = float(agg.echecs   or 0) if agg else 0.0
        abandons_f = float(agg.abandons or 0) if agg else 0.0
        actifs_p   = reussis_f + echecs_f
        return {
            "total":    total_p,
            "reussite": round(reussis_f  / actifs_p * 100, 2) if actifs_p else 0.0,
            "echec":    round(echecs_f   / actifs_p * 100, 2) if actifs_p else 0.0,
            "abandon":  round(abandons_f / total_p  * 100, 2) if total_p  else 0.0,
            "moyenne":  round(float(agg.moy), 2) if agg and agg.moy else 0.0,
        }

    stats_prec_f = calc_stats_annee_f(annee_precedente.id_annee) if annee_precedente else None

    def calc_delta(val_actuelle, val_prec):
        if val_prec is None:
            return None, "stable"
        d = round(float(val_actuelle) - float(val_prec), 1)
        return d, ("up" if d > 0 else "down" if d < 0 else "stable")

    delta_total,    sens_total    = calc_delta(total,            stats_prec_f["total"]    if stats_prec_f else None)
    delta_reussite, sens_reussite = calc_delta(taux_reussite,    stats_prec_f["reussite"] if stats_prec_f else None)
    delta_moyenne,  sens_moyenne  = calc_delta(moyenne_generale, stats_prec_f["moyenne"]  if stats_prec_f else None)
    delta_echec,    sens_echec    = calc_delta(taux_echec,       stats_prec_f["echec"]    if stats_prec_f else None)
    delta_abandon,  sens_abandon  = calc_delta(taux_e,           stats_prec_f["abandon"]  if stats_prec_f else None)

    return render_template(
        "respo/tableau_f.html",
        semestre=semestres, filieres=filieres, niveaux=niveaux,
        semestre_selected=semestre_selected, filiere_selected=filiere_selected,
        niveau_selected=niveau_selected, taux_reussite=taux_reussite,
        taux_echec=taux_echec, taux_e=taux_e, moyenne_generale=moyenne_generale,
        reussite_par_niveau=reussite_par_niveau, echec_par_niveau=echec_par_niveau,
        evolution_par_filiere=evolution_par_filiere,
        annees=[a.libelle for a in annees], annee_active=annee_active,
        total=total, total_actifs=total_actifs,
        nb_critiques_ml=nb_critiques_ml, nb_moderes_ml=nb_moderes_ml,
        nb_abandons_ml=nb_abandons_ml, proba_abandon_moy=proba_abandon_moy,
        title="Filière", annee_precedente=annee_precedente,
        delta_total=delta_total,       sens_total=sens_total,
        delta_reussite=delta_reussite, sens_reussite=sens_reussite,
        delta_moyenne=delta_moyenne,   sens_moyenne=sens_moyenne,
        delta_echec=delta_echec,       sens_echec=sens_echec,
        delta_abandon=delta_abandon,   sens_abandon=sens_abandon,
    )


# ==========================
# TABLEAU MATIÈRES
# ==========================
@app.route('/tableau_m', methods=["GET", "POST"])
@respo_required
def tableau_m():
    annee_active = get_annee_active()

    filiere_selected  = request.args.get("filiere",  "all")
    semestre_selected = request.args.get("semestre", "all")
    niveau_selected   = request.args.get("niveau",   "all")

    filieres  = Filiere.query.order_by(Filiere.nom_filiere).all()
    niveaux   = Niveau.query.order_by(Niveau.libelle).all()
    semestres = get_semestres()

    base_query = (
        db.session.query(Resultat)
        .join(Matiere,     Resultat.id_matiere     == Matiere.id_matiere)
        .join(Inscription, Resultat.id_inscription == Inscription.id_inscription)
        .join(Filiere,     Inscription.id_filiere  == Filiere.id_filiere)
        .join(Niveau,      Inscription.id_niveau   == Niveau.id_niveau)
    )
    if annee_active:
        base_query = base_query.filter(Inscription.id_annee == annee_active.id_annee)

    try:
        if filiere_selected != "all":
            base_query = base_query.filter(Inscription.id_filiere == int(filiere_selected))
        if niveau_selected != "all":
            base_query = base_query.filter(Inscription.id_niveau == int(niveau_selected))
    except ValueError:
        pass

    try:
        if semestre_selected != "all":
            base_query = base_query.filter(Resultat.id_semestre == int(semestre_selected))
    except ValueError:
        pass

    if filiere_selected == "all":
        subq = (
            db.session.query(
                Filiere.id_filiere, Filiere.nom_filiere,
                Matiere.nom_matiere, func.avg(Resultat.moyenne).label("moy")
            )
            .join(Inscription, Resultat.id_inscription == Inscription.id_inscription)
            .join(Filiere,     Inscription.id_filiere  == Filiere.id_filiere)
            .join(Matiere,     Resultat.id_matiere     == Matiere.id_matiere)
        )
        if annee_active:
            subq = subq.filter(Inscription.id_annee == annee_active.id_annee)
        try:
            if niveau_selected != "all":
                subq = subq.filter(Inscription.id_niveau == int(niveau_selected))
        except ValueError:
            pass
        try:
            if semestre_selected != "all":
                subq = subq.filter(Resultat.id_semestre == int(semestre_selected))
        except ValueError:
            pass
        subq = subq.group_by(
            Filiere.id_filiere, Filiere.nom_filiere,
            Matiere.id_matiere, Matiere.nom_matiere
        ).all()

        meilleures = {}
        for id_filiere, nom_filiere, nom_matiere, moy in subq:
            if id_filiere not in meilleures or (moy or 0) > meilleures[id_filiere]["valeur"]:
                meilleures[id_filiere] = {
                    "nom":    f"{nom_filiere} · {nom_matiere}",
                    "valeur": round(moy or 0, 2)
                }
        moyennes = list(meilleures.values())
    else:
        moyennes_agg = (
            base_query
            .with_entities(Matiere.id_matiere, Matiere.nom_matiere, func.avg(Resultat.moyenne))
            .group_by(Matiere.id_matiere, Matiere.nom_matiere)
            .all()
        )
        moyennes = [{"nom": nom, "valeur": round(avg or 0, 2)} for _, nom, avg in moyennes_agg]

    moyennes_par_etudiant = (
        base_query
        .with_entities(Inscription.id_inscription, func.avg(Resultat.moyenne).label("moy_etudiant"))
        .group_by(Inscription.id_inscription)
        .all()
    )
    toutes_moyennes = [r.moy_etudiant for r in moyennes_par_etudiant if r.moy_etudiant is not None]
    distribution = [
        {"intervalle": "0–5",   "effectif": sum(0  <= m <  5  for m in toutes_moyennes)},
        {"intervalle": "5–10",  "effectif": sum(5  <= m < 10  for m in toutes_moyennes)},
        {"intervalle": "10–15", "effectif": sum(10 <= m < 15  for m in toutes_moyennes)},
        {"intervalle": "15–20", "effectif": sum(15 <= m <= 20 for m in toutes_moyennes)},
    ]

    echec_agg = (
        base_query
        .with_entities(
            Matiere.id_matiere,
            Matiere.nom_matiere,
            func.count(func.distinct(Inscription.id_inscription)).label("total"),
            func.count(func.distinct(
                case((Resultat.moyenne < 10, Inscription.id_inscription), else_=None)
            )).label("echecs"),
        )
        .group_by(Matiere.id_matiere, Matiere.nom_matiere)
        .all()
    )

    ids = [row.id_matiere for row in echec_agg]
    matieres_map = {
        m.id_matiere: m
        for m in Matiere.query
            .options(joinedload(Matiere.professeur), joinedload(Matiere.filiere))
            .filter(Matiere.id_matiere.in_(ids))
            .all()
    } if ids else {}

    niveau_query = (
        db.session.query(Matiere.id_matiere, Niveau.libelle)
        .join(Resultat,    Resultat.id_matiere     == Matiere.id_matiere)
        .join(Inscription, Resultat.id_inscription == Inscription.id_inscription)
        .join(Niveau,      Inscription.id_niveau   == Niveau.id_niveau)
    )
    if annee_active:
        niveau_query = niveau_query.filter(Inscription.id_annee == annee_active.id_annee)
    try:
        if filiere_selected != "all":
            niveau_query = niveau_query.filter(Inscription.id_filiere == int(filiere_selected))
        if niveau_selected != "all":
            niveau_query = niveau_query.filter(Inscription.id_niveau == int(niveau_selected))
    except ValueError:
        pass
    try:
        if semestre_selected != "all":
            niveau_query = niveau_query.filter(Resultat.id_semestre == int(semestre_selected))
    except ValueError:
        pass
    niveaux_par_matiere = {}
    for mid, libelle in niveau_query.distinct().all():
        niveaux_par_matiere.setdefault(mid, set()).add(libelle)

    specialite_query = (
        db.session.query(Matiere.id_matiere, Specialite.nom_specialite)
        .join(Resultat,    Resultat.id_matiere      == Matiere.id_matiere)
        .join(Inscription, Resultat.id_inscription  == Inscription.id_inscription)
        .join(Specialite,  Inscription.id_specialite == Specialite.id_specialite)
    )
    if annee_active:
        specialite_query = specialite_query.filter(Inscription.id_annee == annee_active.id_annee)
    try:
        if filiere_selected != "all":
            specialite_query = specialite_query.filter(Inscription.id_filiere == int(filiere_selected))
        if niveau_selected != "all":
            specialite_query = specialite_query.filter(Inscription.id_niveau == int(niveau_selected))
    except ValueError:
        pass
    try:
        if semestre_selected != "all":
            specialite_query = specialite_query.filter(Resultat.id_semestre == int(semestre_selected))
    except ValueError:
        pass
    specialites_par_matiere = {}
    for mid, nom_sp in specialite_query.distinct().all():
        specialites_par_matiere.setdefault(mid, set()).add(nom_sp)

    # ── Comptage ajournés / admis endettés selon semestre ──────────────────
    if semestre_selected == "all":
        # Mode all : on compte les admis (dettes) par matière non validée
        insc_q = _query_inscriptions_avec_relations(
            annee_id=annee_active.id_annee if annee_active else None
        )
        try:
            if filiere_selected != "all":
                insc_q = insc_q.filter(Inscription.id_filiere == int(filiere_selected))
            if niveau_selected != "all":
                insc_q = insc_q.filter(Inscription.id_niveau == int(niveau_selected))
        except ValueError:
            pass
        toutes_insc = insc_q.all()

        ajournes_par_matiere = {}
        for insc in toutes_insc:
            if insc.statut_simple != "Admis (dettes)":
                continue
            for r in insc.resultats:
                if not r.credit_valide and r.matiere:
                    mid = r.id_matiere
                    ajournes_par_matiere[mid] = ajournes_par_matiere.get(mid, 0) + 1

        nb_etudiants_echec_uniques = sum(
            1 for insc in toutes_insc
            if insc.statut_simple == "Admis (dettes)"
        )
    else:
        # S1/S2 : logique SQL existante inchangée
        rows_ajournes_agg = (
            base_query
            .with_entities(
                Matiere.id_matiere,
                func.count(func.distinct(Inscription.id_inscription)).label("nb_ajournes"),
            )
            .filter(Resultat.moyenne < 10)
            .group_by(Matiere.id_matiere)
            .all()
        )
        ajournes_par_matiere = {row.id_matiere: row.nb_ajournes for row in rows_ajournes_agg}

        nb_etudiants_echec_uniques = (
            base_query
            .with_entities(Inscription.id_inscription)
            .filter(Resultat.moyenne < 10)
            .distinct()
            .count()
        )
    # ───────────────────────────────────────────────────────────────────────

    stats = []
    for row in echec_agg:
        total_m     = int(row.total  or 0)
        echecs_m    = float(row.echecs or 0)
        taux_echec  = round((echecs_m / total_m * 100), 2) if total_m else 0
        nb_ajournes = ajournes_par_matiere.get(row.id_matiere, 0)

        if nb_ajournes == 0:
            continue

        m = matieres_map.get(row.id_matiere)
        if not m:
            continue

        niveaux_matiere     = niveaux_par_matiere.get(row.id_matiere, set())
        specialites_matiere = specialites_par_matiere.get(row.id_matiere, set())

        stats.append({
            "id_matiere": row.id_matiere,
            "filiere":    m.filiere.nom_filiere if m.filiere else "-",
            "matiere":    m.nom_matiere,
            "specialite": ", ".join(sorted(specialites_matiere)) if specialites_matiere else "-",
            "professeur": f"{m.professeur.nom} {m.professeur.prenom}" if m.professeur else "-",
            "telephone":  m.professeur.telephone if m.professeur else "-",
            "niveau":     ", ".join(sorted(niveaux_matiere)) if niveaux_matiere else "-",
            "taux_echec": taux_echec,
            "ajournes":   nb_ajournes,
        })

    stats_difficiles = sorted(stats, key=lambda x: x["taux_echec"], reverse=True)

    nb_inscrits_total = (
        base_query
        .with_entities(Inscription.id_inscription)
        .distinct()
        .count()
    )

    return render_template(
        "respo/tableau_m.html",
        filieres=filieres, niveaux=niveaux, semestre=semestres,
        filiere_selected=filiere_selected, semestre_selected=semestre_selected,
        niveau_selected=niveau_selected, moyennes=moyennes,
        distribution=distribution, stats=stats_difficiles,
        nb_matieres=len(stats_difficiles), annee_active=annee_active,
        filiere_mode=(filiere_selected == "all"),
        nb_etudiants_echec=nb_etudiants_echec_uniques,
        nb_inscrits_total=nb_inscrits_total,
        mode_all=(semestre_selected == "all"),
        title="Analyse par matière"
    )


# ==========================
# AJOURNÉS PAR MATIÈRE
# ==========================
@app.route('/ajournes_matiere/<int:id_matiere>')
@respo_required
def ajournes_matiere(id_matiere):
    annee_active = get_annee_active()

    filiere_selected  = request.args.get("filiere",  "all")
    semestre_selected = request.args.get("semestre", "all")
    niveau_selected   = request.args.get("niveau",   "all")

    matiere = (
        Matiere.query
        .options(joinedload(Matiere.professeur), joinedload(Matiere.filiere))
        .get_or_404(id_matiere)
    )

    if semestre_selected == "all":
        # Mode all → admis endettés sur cette matière
        insc_q = _query_inscriptions_avec_relations(
            annee_id=annee_active.id_annee if annee_active else None
        )
        try:
            if filiere_selected != "all":
                insc_q = insc_q.filter(Inscription.id_filiere == int(filiere_selected))
            if niveau_selected != "all":
                insc_q = insc_q.filter(Inscription.id_niveau == int(niveau_selected))
        except ValueError:
            pass

        ajournes = []
        for insc in insc_q.all():
            if insc.statut_simple != "Admis (dettes)":
                continue
            for r in insc.resultats:
                if r.id_matiere == id_matiere and not r.credit_valide:
                    ajournes.append({
                        "matricule":  insc.etudiant.matricule,
                        "nom":        insc.etudiant.nom,
                        "prenom":     insc.etudiant.prenom,
                        "filiere":    insc.filiere.nom_filiere       if insc.filiere    else "—",
                        "specialite": insc.specialite.nom_specialite if insc.specialite else "—",
                        "niveau":     insc.niveau.libelle            if insc.niveau     else "—",
                        "moyenne":    round(float(r.moyenne), 2)     if r.moyenne       else 0,
                    })
                    break
        ajournes.sort(key=lambda x: x["moyenne"])

    else:
        # S1/S2 → logique existante inchangée
        q = (
            db.session.query(
                Etudiant.matricule, Etudiant.nom, Etudiant.prenom,
                Filiere.nom_filiere, Specialite.nom_specialite,
                Niveau.libelle.label("niveau"), Resultat.moyenne
            )
            .join(Inscription, Etudiant.id_etudiant      == Inscription.id_etudiant)
            .join(Resultat,    Inscription.id_inscription == Resultat.id_inscription)
            .join(Filiere,     Inscription.id_filiere     == Filiere.id_filiere)
            .join(Niveau,      Inscription.id_niveau      == Niveau.id_niveau)
            .outerjoin(Specialite, Inscription.id_specialite == Specialite.id_specialite)
            .filter(Resultat.id_matiere == id_matiere, Resultat.moyenne < 10)
        )
        if annee_active:
            q = q.filter(Inscription.id_annee == annee_active.id_annee)
        try:
            if filiere_selected != "all":
                q = q.filter(Inscription.id_filiere == int(filiere_selected))
            if niveau_selected != "all":
                q = q.filter(Inscription.id_niveau == int(niveau_selected))
        except ValueError:
            pass
        try:
            q = q.filter(Resultat.id_semestre == int(semestre_selected))
        except ValueError:
            pass

        rows = q.order_by(Resultat.moyenne.asc()).all()
        ajournes = [
            {
                "matricule":  r.matricule,
                "nom":        r.nom,
                "prenom":     r.prenom,
                "filiere":    r.nom_filiere,
                "specialite": r.nom_specialite or "—",
                "niveau":     r.niveau,
                "moyenne":    round(float(r.moyenne), 2) if r.moyenne is not None else 0,
            }
            for r in rows
        ]

    total_q = (
        db.session.query(func.count(func.distinct(Inscription.id_inscription)))
        .join(Resultat, Inscription.id_inscription == Resultat.id_inscription)
        .filter(Resultat.id_matiere == id_matiere)
    )
    if annee_active:
        total_q = total_q.filter(Inscription.id_annee == annee_active.id_annee)
    total      = total_q.scalar() or 0
    taux_echec = round(len(ajournes) / total * 100, 1) if total else 0

    return render_template(
        "respo/ajournes_matiere.html",
        matiere=matiere, ajournes=ajournes, taux_echec=taux_echec,
        annee_active=annee_active, filiere_selected=filiere_selected,
        semestre_selected=semestre_selected, niveau_selected=niveau_selected,
        mode_all=(semestre_selected == "all"),
        title=f"{'Admis endettés' if semestre_selected == 'all' else 'Ajournés'} — {matiere.nom_matiere}"
    )


@app.route('/ajournes_matiere/<int:id_matiere>/pdf')
@respo_required
def ajournes_matiere_pdf(id_matiere):
    annee_active = get_annee_active()

    filiere_selected  = request.args.get("filiere",  "all")
    semestre_selected = request.args.get("semestre", "all")
    niveau_selected   = request.args.get("niveau",   "all")

    matiere = (
        Matiere.query
        .options(joinedload(Matiere.professeur), joinedload(Matiere.filiere))
        .get_or_404(id_matiere)
    )

    q = (
        db.session.query(
            Etudiant.matricule, Etudiant.nom, Etudiant.prenom,
            Filiere.nom_filiere, Specialite.nom_specialite,
            Niveau.libelle.label("niveau"), Resultat.moyenne
        )
        .join(Inscription, Etudiant.id_etudiant      == Inscription.id_etudiant)
        .join(Resultat,    Inscription.id_inscription == Resultat.id_inscription)
        .join(Filiere,     Inscription.id_filiere     == Filiere.id_filiere)
        .join(Niveau,      Inscription.id_niveau      == Niveau.id_niveau)
        .outerjoin(Specialite, Inscription.id_specialite == Specialite.id_specialite)
        .filter(Resultat.id_matiere == id_matiere, Resultat.moyenne < 10)
    )

    if annee_active:
        q = q.filter(Inscription.id_annee == annee_active.id_annee)
    try:
        if filiere_selected != "all":
            q = q.filter(Inscription.id_filiere == int(filiere_selected))
        if niveau_selected != "all":
            q = q.filter(Inscription.id_niveau == int(niveau_selected))
    except ValueError:
        pass
    try:
        if semestre_selected != "all":
            q = q.filter(Resultat.id_semestre == int(semestre_selected))
    except ValueError:
        pass

    rows = q.order_by(Resultat.moyenne.asc()).all()

    total_q = (
        db.session.query(func.count(func.distinct(Inscription.id_inscription)))
        .join(Resultat, Inscription.id_inscription == Resultat.id_inscription)
        .filter(Resultat.id_matiere == id_matiere)
    )
    if annee_active:
        total_q = total_q.filter(Inscription.id_annee == annee_active.id_annee)
    total       = total_q.scalar() or 0
    nb_ajournes = len(rows)
    taux_echec  = round(nb_ajournes / total * 100, 1) if total else 0

    buffer = BytesIO()
    doc    = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=2*cm,    bottomMargin=2*cm
    )

    C_DARK   = colors.HexColor("#1e2d3d")
    C_BLUE   = colors.HexColor("#3b82f6")
    C_RED    = colors.HexColor("#ef4444")
    C_STRIPE = colors.HexColor("#f7fafd")
    C_WHITE  = colors.white

    s_title  = ParagraphStyle("title",  fontName="Helvetica-Bold",
        fontSize=16, textColor=C_DARK, spaceAfter=6)
    s_sub    = ParagraphStyle("sub",    fontName="Helvetica",
        fontSize=9,  textColor=colors.HexColor("#8da0b8"), spaceAfter=8)
    s_meta   = ParagraphStyle("meta",   fontName="Helvetica",
        fontSize=8.5, textColor=C_DARK)
    s_footer = ParagraphStyle("footer", fontName="Helvetica",
        fontSize=7.5, textColor=colors.HexColor("#8da0b8"), alignment=TA_CENTER)

    story       = []
    annee_lib   = annee_active.libelle if annee_active else "—"
    filiere_lib = matiere.filiere.nom_filiere if matiere.filiere else "—"
    prof_lib    = f"{matiere.professeur.nom} {matiere.professeur.prenom}" if matiere.professeur else "—"
    tel_lib     = matiere.professeur.telephone if matiere.professeur else "—"

    story.append(Paragraph(f"Liste des ajournés — {matiere.nom_matiere}", s_title))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"{filiere_lib}  ·  Année : {annee_lib}", s_sub))
    story.append(HRFlowable(width="100%", thickness=1.5, color=C_BLUE, spaceAfter=8))
    story.append(Paragraph(
        f"<b>Professeur :</b> {prof_lib}   &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Tél :</b> {tel_lib}   &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Ajournés :</b> {nb_ajournes}   &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Taux d'échec :</b> {taux_echec}%", s_meta
    ))
    story.append(Spacer(1, 14))

    header = ["#", "Matricule", "Nom", "Prénom", "Filière", "Spécialité", "Niveau", "Moy./20"]
    data   = [header]
    for i, r in enumerate(rows, 1):
        moy = round(float(r.moyenne), 2) if r.moyenne is not None else 0
        data.append([str(i), r.matricule or "—", r.nom or "—", r.prenom or "—",
                     r.nom_filiere or "—", r.nom_specialite or "—", r.niveau or "—", f"{moy}/20"])

    col_widths = [0.7*cm, 2.4*cm, 3*cm, 3*cm, 2.8*cm, 3*cm, 2*cm, 2*cm]
    tbl_style  = TableStyle([
        ("BACKGROUND",    (0, 0), (-1,  0), C_DARK),
        ("TEXTCOLOR",     (0, 0), (-1,  0), C_WHITE),
        ("FONTNAME",      (0, 0), (-1,  0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1,  0), 7.5),
        ("TOPPADDING",    (0, 0), (-1,  0), 8),
        ("BOTTOMPADDING", (0, 0), (-1,  0), 8),
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 8),
        ("TOPPADDING",    (0, 1), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [C_STRIPE, C_WHITE]),
        ("TEXTCOLOR",     (7, 1), (7, -1), C_RED),
        ("FONTNAME",      (7, 1), (7, -1), "Helvetica-Bold"),
        ("LINEBELOW",     (0, 0), (-1, -1), 0.5, colors.HexColor("#dde5f0")),
        ("ALIGN",         (0, 0), (-1, -1), "LEFT"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
    ])
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(tbl_style)
    story.append(table)
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#dde5f0"), spaceAfter=6))
    story.append(Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — IUAInsight", s_footer))

    doc.build(story)
    buffer.seek(0)

    filename = f"ajournes_{matiere.nom_matiere.replace(' ', '_')}_{annee_lib}.pdf"
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=filename)


# ==========================
# ANALYSE DÉMOGRAPHIQUE
# ==========================
@app.route('/analyse_demo')
@respo_required
def analyse_demo():
    annee_active        = get_annee_active()
    filiere_selected    = request.args.get("filiere",    "all")
    specialite_selected = request.args.get("specialite", "all")
    niveau_selected     = request.args.get("niveau",     "all")

    inscrits_q = (
        db.session.query(Etudiant)
        .join(Inscription)
        .options(joinedload(Etudiant.nationalite))
    )
    if annee_active:
        inscrits_q = inscrits_q.filter(Inscription.id_annee == annee_active.id_annee)
    try:
        if filiere_selected != "all":
            inscrits_q = inscrits_q.filter(Inscription.id_filiere == int(filiere_selected))
        if specialite_selected != "all":
            inscrits_q = inscrits_q.filter(Inscription.id_specialite == int(specialite_selected))
        if niveau_selected != "all":
            inscrits_q = inscrits_q.filter(Inscription.id_niveau == int(niveau_selected))
    except ValueError:
        pass
    etudiants = inscrits_q.distinct().all()

    femmes = sum(1 for e in etudiants if e.genre in ("Féminin", "F"))
    hommes = sum(1 for e in etudiants if e.genre in ("Masculin", "M"))

    ages      = [e.calculer_age() for e in etudiants if e.calculer_age() is not None]
    age_moyen = round(statistics.mean(ages), 1) if ages else 0
    tranches  = [
        {"intervalle": "18-20", "effectif": sum(18 <= a <= 20 for a in ages)},
        {"intervalle": "21-23", "effectif": sum(21 <= a <= 23 for a in ages)},
        {"intervalle": "24-26", "effectif": sum(24 <= a <= 26 for a in ages)},
        {"intervalle": "27+",   "effectif": sum(a >= 27        for a in ages)},
    ]

    nationaux         = sum(1 for e in etudiants if e.nationalite and e.nationalite.pays == "Côte d'Ivoire")
    etrangers         = sum(1 for e in etudiants if e.nationalite and e.nationalite.pays != "Côte d'Ivoire")
    hommes_etrangers  = sum(1 for e in etudiants if e.genre in ("Masculin", "M") and e.nationalite and e.nationalite.pays != "Côte d'Ivoire")
    femmes_etrangeres = sum(1 for e in etudiants if e.genre in ("Féminin",  "F") and e.nationalite and e.nationalite.pays != "Côte d'Ivoire")
    hommes_nationaux  = sum(1 for e in etudiants if e.genre in ("Masculin", "M") and e.nationalite and e.nationalite.pays == "Côte d'Ivoire")
    femmes_nationales = sum(1 for e in etudiants if e.genre in ("Féminin",  "F") and e.nationalite and e.nationalite.pays == "Côte d'Ivoire")
    nb_nationalites   = len({e.nationalite.pays for e in etudiants if e.nationalite and e.nationalite.pays})

    filieres_list = Filiere.query.order_by(Filiere.nom_filiere).all()
    specialites   = Specialite.query.order_by(Specialite.nom_specialite).all()
    niveaux       = Niveau.query.order_by(Niveau.ordre).all()

    agg_filiere_genre = db.session.query(
        Inscription.id_filiere,
        func.lower(func.trim(Etudiant.genre)).label("genre"),
        func.count(func.distinct(Etudiant.id_etudiant)).label("nb")
    ).join(Etudiant, Inscription.id_etudiant == Etudiant.id_etudiant)
    if annee_active:
        agg_filiere_genre = agg_filiere_genre.filter(Inscription.id_annee == annee_active.id_annee)
    agg_filiere_genre = agg_filiere_genre.group_by(
        Inscription.id_filiere, func.lower(func.trim(Etudiant.genre))
    ).all()

    fg_map = {}
    for row in agg_filiere_genre:
        fg_map.setdefault(row.id_filiere, {})
        fg_map[row.id_filiere][row.genre] = row.nb

    filieres_stats = []
    for f in filieres_list:
        genres  = fg_map.get(f.id_filiere, {})
        h       = sum(v for g, v in genres.items() if g in MASCULIN_VALS)
        fe      = sum(v for g, v in genres.items() if g in FEMININ_VALS)
        total_f = h + fe
        if total_f == 0:
            continue
        filieres_stats.append({
            "nom":    f.nom_filiere,
            "hommes": round(h  / total_f * 100, 1),
            "femmes": round(fe / total_f * 100, 1),
        })

    return render_template(
        "respo/analyse_demo.html",
        femmes=femmes, hommes=hommes, age_moyen=age_moyen, ages=tranches,
        nationaux=nationaux, etrangers=etrangers, nb_nationalites=nb_nationalites,
        hommes_etrangers=hommes_etrangers, femmes_etrangeres=femmes_etrangeres,
        hommes_nationaux=hommes_nationaux, femmes_nationales=femmes_nationales,
        filieres=filieres_stats, filieres_list=filieres_list,
        specialites=specialites, niveaux=niveaux,
        filiere_selected=filiere_selected, specialite_selected=specialite_selected,
        niveau_selected=niveau_selected, annee_active=annee_active,
        title='Analyse démographique'
    )


# ==========================
# RAPPORT
# ==========================
@app.route('/rapport', methods=["GET", "POST"])
@respo_required
def rapport():
    form         = RapportForm()
    annee_active = get_annee_active()
    annees_list  = AnneeScolaire.query.order_by(AnneeScolaire.date_debut.desc()).all()
    filieres     = Filiere.query.order_by(Filiere.nom_filiere).all()
    niveaux      = Niveau.query.order_by(Niveau.libelle).all()

    if annee_active:
        form.periode.choices = [
            (f"{annee_active.libelle} · Annuel",     f"{annee_active.libelle} · Annuel"),
            (f"{annee_active.libelle} · Semestre 1", f"{annee_active.libelle} · Semestre 1"),
            (f"{annee_active.libelle} · Semestre 2", f"{annee_active.libelle} · Semestre 2"),
        ]
    else:
        form.periode.choices = [("", "Aucune année active")]

    form.filiere.choices = [("Toutes", "Toutes les filières")] + [(f.nom_filiere, f.nom_filiere) for f in filieres]
    form.niveau.choices  = [("Tous",   "Tous les niveaux")]    + [(n.libelle,     n.libelle)     for n in niveaux]
    periode_defaut = f"{annee_active.libelle} · Annuel" if annee_active else ""

    if form.validate_on_submit():
        type_rapport   = form.type_rapport.data
        periode        = form.periode.data
        filiere_nom    = form.filiere.data
        niveau_nom     = form.niveau.data
        format_rapport = form.format_fichier.data

        libelle_choisi = periode.split("·")[0].strip()
        annee_choisie  = AnneeScolaire.query.filter_by(libelle=libelle_choisi).first()

        def base_q():
            q = _query_inscriptions_avec_relations(
                annee_id=annee_choisie.id_annee if annee_choisie else None
            ).join(Etudiant, Inscription.id_etudiant == Etudiant.id_etudiant)\
             .join(Filiere,  Inscription.id_filiere  == Filiere.id_filiere)\
             .join(Niveau,   Inscription.id_niveau   == Niveau.id_niveau)
            if filiere_nom != "Toutes":
                q = q.filter(Filiere.nom_filiere == filiere_nom)
            if niveau_nom != "Tous":
                q = q.filter(Niveau.libelle == niveau_nom)
            return q

        if type_rapport == "Rapport de résultats":
            q = base_q()
            if "Semestre 1" in periode:
                inscriptions = q.filter(Inscription.moyenne_s1.isnot(None)).all()
                moyennes     = [i.moyenne_s1 for i in inscriptions if i.moyenne_s1 is not None]
                abandon      = 0
            elif "Semestre 2" in periode:
                inscriptions = q.filter(Inscription.moyenne_s2.isnot(None)).all()
                moyennes     = [i.moyenne_s2 for i in inscriptions if i.moyenne_s2 is not None]
                abandon      = 0
            else:
                inscriptions = q.filter(Inscription.moyenne_annuelle.isnot(None)).all()
                moyennes     = [i.moyenne_annuelle for i in inscriptions if i.moyenne_annuelle is not None]
                abandon      = base_q().filter(
                    Inscription.moyenne_s1.is_(None), Inscription.moyenne_s2.is_(None)
                ).count()
            total  = len(inscriptions)
            admis  = sum(1 for m in moyennes if m >= 10)
            echecs = sum(1 for m in moyennes if m < 10)
            tx_r   = round((admis  / total * 100), 1) if total else 0
            tx_e   = round((echecs / total * 100), 1) if total else 0
            ml_results_rapport = predict_batch(inscriptions)
            nb_critiques_r = sum(1 for r in ml_results_rapport if r and r["niveau_risque"] == "critique")
            nb_abandons_r  = sum(1 for r in ml_results_rapport if r and r["probabilite_abandon"] >= 0.7)
            if "Annuel" in periode:
                tx_a     = round((abandon / (total + abandon) * 100), 1) if (total + abandon) else 0
                kpi_rows = [
                    ["Indicateur", "Valeur"], ["Total inscrits", str(total + abandon)],
                    ["Admis", str(admis)], ["Échecs", str(echecs)], ["Abandons", str(abandon)],
                    ["Taux de réussite", f"{tx_r} %"], ["Taux d'échec", f"{tx_e} %"], ["Taux d'abandon", f"{tx_a} %"],
                    ["ML — Risque critique", str(nb_critiques_r)], ["ML — Risque abandon (≥70%)", str(nb_abandons_r)],
                ]
            else:
                kpi_rows = [
                    ["Indicateur", "Valeur"], ["Total inscrits", str(total)],
                    ["Admis", str(admis)], ["Échecs", str(echecs)],
                    ["Taux de réussite", f"{tx_r} %"], ["Taux d'échec", f"{tx_e} %"],
                    ["ML — Risque critique", str(nb_critiques_r)], ["ML — Risque abandon (≥70%)", str(nb_abandons_r)],
                ]

        elif type_rapport == "Étudiants admis":
            inscriptions  = base_q().filter(
                Inscription.moyenne_annuelle.isnot(None), Inscription.moyenne_annuelle >= 10
            ).all()
            total_admis   = len(inscriptions)
            total_general = base_q().filter(Inscription.moyenne_annuelle.isnot(None)).count()
            tx_r = round((total_admis / total_general * 100), 1) if total_general else 0
            kpi_rows = [["Indicateur", "Valeur"], ["Total admis", str(total_admis)], ["Taux de réussite", f"{tx_r} %"]]

        elif type_rapport == "Étudiants en échec":
            inscriptions  = base_q().filter(
                Inscription.moyenne_annuelle.isnot(None), Inscription.moyenne_annuelle < 10
            ).all()
            total_echec   = len(inscriptions)
            total_general = base_q().filter(Inscription.moyenne_annuelle.isnot(None)).count()
            tx_e = round((total_echec / total_general * 100), 1) if total_general else 0
            ml_results_echec = predict_batch(inscriptions)
            nb_critiques_e = sum(1 for r in ml_results_echec if r and r["niveau_risque"] == "critique")
            kpi_rows = [
                ["Indicateur", "Valeur"], ["Total en échec", str(total_echec)],
                ["Taux d'échec", f"{tx_e} %"], ["ML — Risque critique parmi échecs", str(nb_critiques_e)],
            ]

        elif type_rapport == "Abandon étudiants":
            inscriptions = base_q().filter(
                Inscription.moyenne_s1.is_(None), Inscription.moyenne_s2.is_(None)
            ).all()
            total = len(inscriptions)
            ml_results_abandon = predict_batch(inscriptions)
            nb_abandon_ml = sum(1 for r in ml_results_abandon if r and r["probabilite_abandon"] >= 0.7)
            kpi_rows = [
                ["Indicateur", "Valeur"], ["Total abandons", str(total)],
                ["ML — Abandon probable confirmé", str(nb_abandon_ml)],
            ]

        else:
            inscriptions = base_q().all()
            total  = len(inscriptions)
            admis  = sum(1 for i in inscriptions if i.moyenne_annuelle and i.moyenne_annuelle >= 10)
            echecs = sum(1 for i in inscriptions if i.moyenne_annuelle and i.moyenne_annuelle < 10)
            tx_r   = round((admis  / total * 100), 1) if total else 0
            tx_e   = round((echecs / total * 100), 1) if total else 0
            ml_results_all   = predict_batch(inscriptions)
            nb_critiques_all = sum(1 for r in ml_results_all if r and r["niveau_risque"] == "critique")
            kpi_rows = [
                ["Indicateur", "Valeur"], ["Total inscrits", str(total)],
                ["Admis", str(admis)], ["Échecs", str(echecs)],
                ["Taux de réussite", f"{tx_r} %"], ["Taux d'échec", f"{tx_e} %"],
                ["Risque critique", str(nb_critiques_all)],
            ]

        if type_rapport == "Rapport de résultats":
            if "Semestre 1" in periode:
                headers = ["Matricule", "Nom", "Prénom", "Filière", "Niveau", "Moy. S1", "Statut", "Risque", "Abandon"]
            elif "Semestre 2" in periode:
                headers = ["Matricule", "Nom", "Prénom", "Filière", "Niveau", "Moy. S2", "Statut", "Risque", "Abandon"]
            else:
                headers = ["Matricule", "Nom", "Prénom", "Filière", "Niveau", "Moy. S1", "Moy. S2", "Moy. Ann.", "Statut", "Risque ML", "Abandon ML"]
        elif type_rapport in ("Étudiants admis", "Étudiants en échec"):
            headers = ["Matricule", "Nom", "Prénom", "Filière", "Niveau", "Moy. Ann.", "Statut", "Risque", "Abandon"]
        elif type_rapport == "Abandon étudiants":
            headers = ["Matricule", "Nom", "Prénom", "Filière", "Niveau", "Statut", "Risque ML", "P(Abandon)"]
        else:
            headers = ["Matricule", "Nom", "Prénom", "Filière", "Niveau", "Moy. S1", "Moy. S2", "Moy. Ann.", "Statut", "Risque", "Abandon"]

        ml_results_final_list = predict_batch(inscriptions)
        ml_results_final      = ml_index(inscriptions, ml_results_final_list)

        archives_dir = os.path.join(app.root_path, "static", "archives")
        os.makedirs(archives_dir, exist_ok=True)
        safe_titre  = type_rapport.replace(" ", "_").replace("/", "-")
        safe_period = periode.replace(" ", "_").replace("·", "-").replace("/", "-")
        ext         = "pdf" if format_rapport == "PDF" else "xlsx"
        filename    = f"{safe_titre}_{safe_period}.{ext}"
        filepath    = os.path.join(archives_dir, filename)

        def build_row(i):
            fil       = i.filiere.nom_filiere if i.filiere else "-"
            niv       = i.niveau.libelle      if i.niveau  else "-"
            ml_r      = ml_results_final.get(i.id_inscription)
            ml_risque  = ml_r["niveau_risque"]                     if ml_r else "-"
            ml_abandon = f"{ml_r['probabilite_abandon']*100:.0f}%" if ml_r else "-"
            if type_rapport == "Abandon étudiants":
                return [i.etudiant.matricule, i.etudiant.nom, i.etudiant.prenom, fil, niv, "Abandon", ml_risque, ml_abandon]
            elif type_rapport == "Étudiants admis":
                return [i.etudiant.matricule, i.etudiant.nom, i.etudiant.prenom, fil, niv,
                        str(i.moyenne_annuelle or "-"), "Admis", ml_risque, ml_abandon]
            elif type_rapport == "Étudiants en échec":
                return [i.etudiant.matricule, i.etudiant.nom, i.etudiant.prenom, fil, niv,
                        str(i.moyenne_annuelle or "-"), "Échec", ml_risque, ml_abandon]
            elif type_rapport == "Rapport de résultats":
                if "Semestre 1" in periode:
                    moy    = i.moyenne_s1
                    statut = "Admis" if moy and moy >= 10 else ("Échec" if moy else "—")
                    return [i.etudiant.matricule, i.etudiant.nom, i.etudiant.prenom, fil, niv,
                            str(moy or "-"), statut, ml_risque, ml_abandon]
                elif "Semestre 2" in periode:
                    moy    = i.moyenne_s2
                    statut = "Admis" if moy and moy >= 10 else ("Échec" if moy else "—")
                    return [i.etudiant.matricule, i.etudiant.nom, i.etudiant.prenom, fil, niv,
                            str(moy or "-"), statut, ml_risque, ml_abandon]
                else:
                    statut = "Admis" if i.moyenne_annuelle and i.moyenne_annuelle >= 10 else ("Échec" if i.moyenne_annuelle else "—")
                    return [i.etudiant.matricule, i.etudiant.nom, i.etudiant.prenom, fil, niv,
                            str(i.moyenne_s1 or "-"), str(i.moyenne_s2 or "-"), str(i.moyenne_annuelle or "-"),
                            statut, ml_risque, ml_abandon]
            else:
                statut = "Admis" if i.moyenne_annuelle and i.moyenne_annuelle >= 10 else ("Échec" if i.moyenne_annuelle else "—")
                return [i.etudiant.matricule, i.etudiant.nom, i.etudiant.prenom, fil, niv,
                        str(i.moyenne_s1 or "-"), str(i.moyenne_s2 or "-"), str(i.moyenne_annuelle or "-"),
                        statut, ml_risque, ml_abandon]

        if format_rapport == "PDF":
            doc    = SimpleDocTemplate(filepath, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
            styles = getSampleStyleSheet()
            story  = []
            story.append(Paragraph("IUADECIS — Rapport Académique",
                ParagraphStyle("titre", parent=styles["Title"], fontSize=18, textColor=colors.HexColor("#1e2d3d"), spaceAfter=4)))
            story.append(Paragraph(f"{type_rapport} &bull; {periode}",
                ParagraphStyle("sous", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#8da0b8"), spaceAfter=4)))
            story.append(Paragraph(f"Filière : {filiere_nom} &nbsp;|&nbsp; Niveau : {niveau_nom}",
                ParagraphStyle("sous2", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#8da0b8"), spaceAfter=10)))
            story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e8eef6"), spaceAfter=14))
            kpi_table = Table(kpi_rows, colWidths=[8*cm, 6*cm])
            kpi_table.setStyle(TableStyle([
                ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#1e2d3d")),
                ("TEXTCOLOR",  (0,0),(-1,0), colors.white),
                ("FONTNAME",   (0,0),(-1,0), "Helvetica-Bold"),
                ("FONTSIZE",   (0,0),(-1,-1), 9),
                ("ROWBACKGROUNDS", (0,1),(-1,-1), [colors.HexColor("#f7f9fc"), colors.white]),
                ("GRID",       (0,0),(-1,-1), 0.5, colors.HexColor("#e8eef6")),
                ("LEFTPADDING",(0,0),(-1,-1), 8),
                ("TOPPADDING", (0,0),(-1,-1), 5),
                ("BOTTOMPADDING",(0,0),(-1,-1), 5),
            ]))
            story.append(kpi_table)
            story.append(Spacer(1, 18))
            story.append(Paragraph("Détail des étudiants", styles["Heading2"]))
            story.append(Spacer(1, 6))
            nb_cols    = len(headers)
            col_width  = 17 * cm / nb_cols
            data       = [headers] + [build_row(i) for i in inscriptions]
            etud_table = Table(data, colWidths=[col_width]*nb_cols, repeatRows=1)
            etud_table.setStyle(TableStyle([
                ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#3a57a7")),
                ("TEXTCOLOR",  (0,0),(-1,0), colors.white),
                ("FONTNAME",   (0,0),(-1,0), "Helvetica-Bold"),
                ("FONTSIZE",   (0,0),(-1,-1), 7.5),
                ("ROWBACKGROUNDS", (0,1),(-1,-1), [colors.HexColor("#f7f9fc"), colors.white]),
                ("GRID",       (0,0),(-1,-1), 0.4, colors.HexColor("#e8eef6")),
                ("LEFTPADDING",(0,0),(-1,-1), 4),
                ("TOPPADDING", (0,0),(-1,-1), 4),
                ("BOTTOMPADDING",(0,0),(-1,-1), 4),
            ]))
            story.append(etud_table)
            story.append(Spacer(1, 20))
            story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e8eef6")))
            story.append(Paragraph(
                f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — IUADECIS · Analyse : {ml_engine.status().get('risk_model_trained') and 'Modèle entraîné' or 'Heuristique'}",
                ParagraphStyle("footer", parent=styles["Normal"], fontSize=7, textColor=colors.HexColor("#aab4c4"), alignment=1)
            ))
            doc.build(story)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Rapport"
            h_font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            h_fill_dark = PatternFill("solid", fgColor="1e2d3d")
            h_fill_blue = PatternFill("solid", fgColor="3a57a7")
            alt_fill    = PatternFill("solid", fgColor="F7F9FC")
            bd          = Side(style="thin", color="E8EEF6")
            cell_border = Border(left=bd, right=bd, top=bd, bottom=bd)
            center      = Alignment(horizontal="center", vertical="center")
            nb_cols_excel = len(headers)
            last_col      = get_column_letter(nb_cols_excel)
            ws.merge_cells(f"A1:{last_col}1")
            ws["A1"] = "IUADECIS — Rapport Académique"
            ws["A1"].font      = Font(name="Arial", bold=True, size=14, color="1e2d3d")
            ws["A1"].alignment = center
            ws.row_dimensions[1].height = 28
            ws.merge_cells(f"A2:{last_col}2")
            ws["A2"] = f"{type_rapport} · {periode} | Filière : {filiere_nom} | Niveau : {niveau_nom}"
            ws["A2"].font      = Font(name="Arial", size=9, color="8da0b8")
            ws["A2"].alignment = center
            ws.row_dimensions[2].height = 16
            for col, h in enumerate(["Indicateur", "Valeur"], start=1):
                c = ws.cell(row=4, column=col, value=h)
                c.font = h_font; c.fill = h_fill_dark; c.alignment = center; c.border = cell_border
            for r, row in enumerate(kpi_rows[1:], start=5):
                for col, v in enumerate(row, start=1):
                    c = ws.cell(row=r, column=col, value=v)
                    c.font = Font(name="Arial", size=9); c.border = cell_border
                    c.alignment = Alignment(vertical="center")
                    if r % 2 == 0: c.fill = alt_fill
            H_ROW = 5 + len(kpi_rows)
            for col, h in enumerate(headers, start=1):
                c = ws.cell(row=H_ROW, column=col, value=h)
                c.font = h_font; c.fill = h_fill_blue; c.alignment = center; c.border = cell_border
            ws.row_dimensions[H_ROW].height = 20
            for r, i in enumerate(inscriptions, start=H_ROW + 1):
                row_vals = build_row(i)
                for col, val in enumerate(row_vals, start=1):
                    c = ws.cell(row=r, column=col, value=val)
                    c.font = Font(name="Arial", size=9); c.border = cell_border
                    c.alignment = Alignment(vertical="center")
                    if r % 2 == 0: c.fill = alt_fill
                sc = ws.cell(row=r, column=len(headers) - 1)
                lv = row_vals[-2]
                sc.font = Font(name="Arial", size=9, bold=True,
                    color=("c0392b" if lv == "critique" else "f59e0b" if lv == "modere" else "1a7a4a" if lv == "ok" else "8da0b8"))
            col_widths_xlsx = {1:14, 2:16, 3:16, 4:20, 5:10, 6:10, 7:10, 8:14, 9:12, 10:14, 11:14}
            for col in range(1, nb_cols_excel + 1):
                ws.column_dimensions[get_column_letter(col)].width = col_widths_xlsx.get(col, 12)
            ws.freeze_panes = f"A{H_ROW + 1}"
            wb.save(filepath)

        rapport_obj = Rapport(
            titre   = type_rapport,
            details = f"{periode} · {filiere_nom} · {niveau_nom}",
            date    = datetime.now(timezone.utc),
            format  = format_rapport,
            path    = filename,
        )
        db.session.add(rapport_obj)
        if type_rapport == "Abandon étudiants" and total > 0:
            upsert_alerte("rapport_abandon", f"Rapport abandon : {total} étudiant(s) · {periode}")
        db.session.commit()
        return redirect(url_for('rapport', dl=filename))

    rapports = Rapport.query.order_by(Rapport.date.desc()).limit(10).all()
    return render_template(
        "respo/rapport.html",
        form=form, annees_list=annees_list, filieres=filieres,
        niveaux=niveaux, rapports=rapports, annee_active=annee_active,
        periode_defaut=periode_defaut, title='Rapport'
    )


@app.route("/rapport/delete/<int:rapport_id>", methods=["POST"])
@respo_required
def rapport_delete(rapport_id):
    rapport = db.get_or_404(Rapport, rapport_id)
    try:
        db.session.delete(rapport); db.session.commit()
        flash("Rapport supprimé avec succès.", "success")
    except Exception:
        db.session.rollback()
        flash("Erreur lors de la suppression du rapport.", "danger")
    return redirect(url_for("rapport"))


@app.route("/rapport/clear", methods=["POST"])
@respo_required
def rapport_clear():
    try:
        Rapport.query.delete(); db.session.commit()
        flash("Historique des rapports vidé avec succès.", "success")
    except Exception:
        db.session.rollback()
        flash("Erreur lors de la suppression de l'historique.", "danger")
    return redirect(url_for("rapport"))


@app.route('/rapport_download/<filename>')
@login_required
def rapport_download(filename):
    archives_dir = os.path.join(app.root_path, "static", "archives")
    filepath     = os.path.join(archives_dir, filename)
    if not os.path.exists(filepath):
        abort(404)
    return send_file(filepath, as_attachment=True)


# ==========================
# TENDANCES
# ==========================
@app.route("/tendances")
@respo_required
def tendances():
    annee_active = get_annee_active()
    annees = list(reversed(AnneeScolaire.query.order_by(AnneeScolaire.libelle.desc()).limit(5).all()))

    inscriptions_evolution = []
    for a in annees:
        nb = db.session.query(func.count(Inscription.id_inscription))\
            .filter(Inscription.id_annee == a.id_annee).scalar()
        inscriptions_evolution.append({"annee": a.libelle, "nombre": nb or 0})

    filieres  = Filiere.query.order_by(Filiere.nom_filiere).all()
    annee_ids = [a.id_annee for a in annees]
    agg_rows  = db.session.query(
        Inscription.id_filiere, Inscription.id_annee,
        func.count(Inscription.id_inscription).label("total"),
        func.sum(case((Inscription.moyenne_annuelle >= 10, 1), else_=0)).label("admis"),
        func.sum(case((Inscription.moyenne_annuelle <  10, 1), else_=0)).label("echecs"),
    ).filter(Inscription.id_annee.in_(annee_ids))\
     .group_by(Inscription.id_filiere, Inscription.id_annee).all()

    agg_index        = {(row.id_filiere, row.id_annee): row for row in agg_rows}
    filiere_reussite = {}
    filiere_echec    = {}
    for f in filieres:
        filiere_reussite[f.nom_filiere] = []
        filiere_echec[f.nom_filiere]    = []
        for a in annees:
            row    = agg_index.get((f.id_filiere, a.id_annee))
            total  = int(row.total  if row else 0)
            admis  = float(row.admis  if row else 0)
            echecs = float(row.echecs if row else 0)
            filiere_reussite[f.nom_filiere].append(round((admis  / total * 100), 2) if total else 0)
            filiere_echec[f.nom_filiere].append(   round((echecs / total * 100), 2) if total else 0)

    global_agg = db.session.query(
        Inscription.id_annee,
        func.count(Inscription.id_inscription).label("total"),
        func.sum(case((Inscription.moyenne_annuelle >= 10, 1), else_=0)).label("admis"),
        func.sum(case((Inscription.moyenne_annuelle <  10, 1), else_=0)).label("echecs"),
    ).filter(Inscription.id_annee.in_(annee_ids)).group_by(Inscription.id_annee).all()

    global_index         = {row.id_annee: row for row in global_agg}
    taux_reussite_annuel = []
    taux_echec_annuel    = []
    for a in annees:
        row    = global_index.get(a.id_annee)
        total  = int(row.total  if row else 0)
        admis  = float(row.admis  if row else 0)
        echecs = float(row.echecs if row else 0)
        taux_reussite_annuel.append(round((admis  / total * 100), 2) if total else 0)
        taux_echec_annuel.append(   round((echecs / total * 100), 2) if total else 0)

    return render_template(
        "respo/tendances.html",
        inscriptions=inscriptions_evolution, annees=[a.libelle for a in annees],
        filieres=filiere_reussite, taux_reussite_annuel=taux_reussite_annuel,
        taux_echec_annuel=taux_echec_annuel, filiere_reussite=filiere_reussite,
        filiere_echec=filiere_echec, annee_active=annee_active, title='Tendances'
    )


# ==========================
# ÉTUDIANTS
# ==========================
@app.route('/student')
@respo_required
def student():
    annee_active = get_annee_active()

    filiere_selected    = request.values.get("filiere",    "all")
    specialite_selected = request.values.get("specialite", "all")
    niveau_selected     = request.values.get("niveau",     "all")
    semestre_selected   = request.values.get("semestre",   "all")
    q = request.args.get("q", "").strip().lower()

    filieres    = Filiere.query.order_by(Filiere.nom_filiere).all()
    specialites = Specialite.query.order_by(Specialite.nom_specialite).all()
    niveaux     = Niveau.query.order_by(Niveau.libelle).all()
    semestres   = get_semestres()

    # ── Dictionnaire spécialités groupées par filière (pour le filtre dynamique) ──
    specialites_par_filiere = {}
    for s in specialites:
        fid = str(s.id_filiere)
        if fid not in specialites_par_filiere:
            specialites_par_filiere[fid] = []
        specialites_par_filiere[fid].append({
            'id':  s.id_specialite,
            'nom': s.nom_specialite
        })

    query = _query_inscriptions_avec_relations(
        annee_id=annee_active.id_annee if annee_active else None
    ).join(Etudiant).join(Filiere).join(Niveau)

    try:
        if filiere_selected    != "all": query = query.filter(Inscription.id_filiere    == int(filiere_selected))
        if specialite_selected != "all": query = query.filter(Inscription.id_specialite == int(specialite_selected))
        if niveau_selected     != "all": query = query.filter(Inscription.id_niveau     == int(niveau_selected))
    except ValueError:
        pass
    query = appliquer_filtre_semestre(query, semestre_selected)
    if q:
        query = query.filter(or_(
            Etudiant.nom.ilike(f"%{q}%"),
            Etudiant.prenom.ilike(f"%{q}%"),
            Etudiant.matricule.ilike(f"%{q}%"),
        ))
    inscriptions = query.all()

    ml_results  = predict_batch(inscriptions)
    ml_par_insc = ml_index(inscriptions, ml_results)

    liste = []
    for i, ml_result in zip(inscriptions, ml_results):
        statut           = i.statut_simple
        credits_restants = max(0, (i.niveau.credits_requis if i.niveau else 60) - (i.credits_valides or 0))
        ml_label, ml_niv_css             = ml_badge(ml_result)
        ml_abandon_label, ml_abandon_css = ml_badge_abandon(ml_result)
        ml_score                         = score_ml_to_int(ml_result)

        liste.append({
            "id":               i.id_inscription,
            "matricule":        i.etudiant.matricule,
            "nom":              i.etudiant.nom,
            "prenom":           i.etudiant.prenom,
            "filiere":          i.filiere.nom_filiere       if i.filiere    else "-",
            "specialite":       i.specialite.nom_specialite if i.specialite else "-",
            "niveau":           i.niveau.libelle            if i.niveau     else "-",
            "moyenne_s1":       i.moyenne_s1       or "-",
            "moyenne_s2":       i.moyenne_s2       or "-",
            "moyenne_annuelle": i.moyenne_annuelle or "-",
            "statut":           statut,
            "credits_valides":  i.credits_valides,
            "credits_restants": credits_restants,
            "score_risque":        ml_score,
            "risque_label":        ml_label,
            "risque_niv":          ml_niv_css,
            "ml_confiance":        ml_result["confiance_risque"]   if ml_result else None,
            "ml_proba_abandon":    ml_result["probabilite_abandon"] if ml_result else None,
            "ml_niveau_abandon":   ml_abandon_label,
            "ml_abandon_css":      ml_abandon_css,
            "ml_source":           ml_result["source"]              if ml_result else "N/A",
            "ml_result":           ml_result,
        })

    liste.sort(key=lambda x: x["score_risque"], reverse=True)

    total    = len(inscriptions)
    admis    = sum(1 for i in inscriptions if i.statut_simple in ("Admis", "Admis (dettes)"))
    echec    = sum(1 for i in inscriptions if i.statut_simple == "Redoublant")
    en_cours = sum(1 for i in inscriptions if i.statut_simple in ("En cours", "Ajourné S1", "Ajourné S2"))
    abandon  = sum(1 for i in inscriptions if i.statut_simple == "Abandon")

    scores         = [e["score_risque"] for e in liste]
    score_moyen    = round(sum(scores) / len(scores), 1) if scores else 0
    nb_critiques   = sum(1 for e in liste if e["risque_niv"] == "critique")
    nb_moderes     = sum(1 for e in liste if e["risque_niv"] == "modere")
    nb_abandons_ml = sum(1 for e in liste if e.get("ml_proba_abandon") and e["ml_proba_abandon"] >= 0.7)

    return render_template(
        "respo/student.html",
        liste=liste, total=total, admis=admis, echec=echec,
        en_cours=en_cours, abandon=abandon, inscrits=total,
        filieres=filieres, specialites=specialites, niveaux=niveaux,
        semestre=semestres, filiere_selected=filiere_selected,
        specialite_selected=specialite_selected, niveau_selected=niveau_selected,
        semestre_selected=semestre_selected, annee_active=annee_active,
        score_moyen=score_moyen, nb_critiques=nb_critiques,
        nb_moderes=nb_moderes, nb_abandons_ml=nb_abandons_ml,
        specialites_par_filiere=specialites_par_filiere,
        ml_source=ml_engine.status().get("risk_model_trained") and "ml" or "heuristic",
        title='Etudiants'
    )
# ==========================
# NOTES ÉTUDIANT
# ==========================
@app.route('/etudiant/<matricule>/notes')
@respo_required
def notes_etudiant(matricule):
    semestre_selected = request.args.get('semestre', 'all')
    etudiant = Etudiant.query.filter_by(matricule=matricule).first_or_404()

    inscription = (
        Inscription.query
        .filter_by(id_etudiant=etudiant.id_etudiant)
        .options(
            joinedload(Inscription.niveau).joinedload(Niveau.niveau_suivant),
            joinedload(Inscription.niveau).selectinload(Niveau.semestres),  # ← FIX
            joinedload(Inscription.filiere),
            selectinload(Inscription.resultats).joinedload(Resultat.matiere),
            selectinload(Inscription.resultats).joinedload(Resultat.semestre),
            selectinload(Inscription.notes).joinedload(Note.session),
        )
        .join(AnneeScolaire, Inscription.id_annee == AnneeScolaire.id_annee)
        .filter(AnneeScolaire.active == True).first()
    )
    if not inscription:
        inscription = (
            Inscription.query
            .filter_by(id_etudiant=etudiant.id_etudiant)
            .options(
                joinedload(Inscription.niveau).joinedload(Niveau.niveau_suivant),
                joinedload(Inscription.niveau).selectinload(Niveau.semestres),  # ← FIX
                joinedload(Inscription.filiere),
                selectinload(Inscription.resultats).joinedload(Resultat.matiere),
                selectinload(Inscription.resultats).joinedload(Resultat.semestre),
                selectinload(Inscription.notes).joinedload(Note.session),
            )
            .order_by(Inscription.id_inscription.desc()).first()
        )
    if not inscription:
        return "Inscription introuvable", 404

    toutes_inscriptions = Inscription.query.filter_by(id_etudiant=etudiant.id_etudiant).all()
    ids_inscriptions    = [i.id_inscription for i in toutes_inscriptions]

    semestres_dispo = Semestre.query\
        .join(Resultat, Resultat.id_semestre == Semestre.id_semestre)\
        .filter(Resultat.id_inscription.in_(ids_inscriptions))\
        .distinct().order_by(Semestre.id_semestre).all()

    query = (
        Resultat.query
        .filter(Resultat.id_inscription.in_(ids_inscriptions))
        .options(joinedload(Resultat.matiere), joinedload(Resultat.semestre))
    )

    semestre_id_filtre = None
    if semestre_selected != 'all':
        try:
            semestre_id_filtre = int(semestre_selected)
            query = query.filter(Resultat.id_semestre == semestre_id_filtre)
        except ValueError:
            pass

    resultats = query.all()

    notes_all = (
        Note.query
        .filter(Note.id_inscription.in_(ids_inscriptions))
        .options(joinedload(Note.session))
        .all()
    )
    notes_index = {}
    for n in notes_all:
        key = (n.id_inscription, n.id_matiere)
        notes_index.setdefault(key, []).append(n)

    resultats_par_semestre = {}
    for r in resultats:
        sem_label = r.semestre.libelle
        if sem_label not in resultats_par_semestre:
            resultats_par_semestre[sem_label] = []

        notes_raw      = notes_index.get((r.id_inscription, r.id_matiere), [])
        notes_groupees = {}
        for n in notes_raw:
            sl = n.session.libelle
            if sl not in notes_groupees:
                notes_groupees[sl] = {}
            notes_groupees[sl][n.type_evaluation] = n.valeur

        resultats_par_semestre[sem_label].append({
            "nom":           r.matiere.nom_matiere,
            "credits":       r.matiere.credit,
            "moyenne":       r.moyenne,
            "credit_valide": r.credit_valide,
            "notes":         notes_groupees,
        })

    statut         = inscription.statut_simple
    tous_resultats = inscription.resultats
    resultats_scope = [r for r in tous_resultats if r.id_semestre == semestre_id_filtre] if semestre_id_filtre else tous_resultats
    credits_valides = sum(r.matiere.credit for r in resultats_scope if r.credit_valide and r.matiere)
    credits_total   = (
        sum(r.matiere.credit for r in resultats_scope if r.matiere)
        if semestre_id_filtre
        else (inscription.niveau.credits_requis if inscription.niveau else 60)
    )

    if semestre_id_filtre is not None:
        sem_obj = Semestre.query.get(semestre_id_filtre)
        if sem_obj and hasattr(sem_obj, 'ordre'):
            moyenne_affichee = inscription.moyenne_s1 if sem_obj.ordre % 2 == 1 else inscription.moyenne_s2
        else:
            moys = [r.moyenne for r in resultats_scope if r.moyenne is not None]
            moyenne_affichee = round(sum(moys) / len(moys), 2) if moys else None
        label_moyenne = f"Moy. {sem_obj.libelle}" if sem_obj else "Moyenne"
    elif inscription.moyenne_s1 is not None and inscription.moyenne_s2 is not None:
        moyenne_affichee = round((inscription.moyenne_s1 + inscription.moyenne_s2) / 2, 2)
        label_moyenne    = "Moyenne annuelle"
    elif inscription.moyenne_s1 is not None:
        moyenne_affichee = inscription.moyenne_s1
        label_moyenne    = "Moy. S1"
    else:
        moyenne_affichee = inscription.moyenne_annuelle
        label_moyenne    = "Moyenne"

    ml_result                        = predict_ml(inscription)
    ml_label, ml_niv                 = ml_badge(ml_result)
    ml_abandon_label, ml_abandon_css = ml_badge_abandon(ml_result)

    return render_template(
        'respo/notes_etudiant.html',
        etudiant=etudiant, inscription=inscription,
        semestres=semestres_dispo, semestre_selected=semestre_selected,
        resultats_par_semestre=resultats_par_semestre,
        statut=statut, credits_valides=credits_valides, credits_total=credits_total,
        moyenne_affichee=moyenne_affichee, label_moyenne=label_moyenne,
        score_risque=score_ml_to_int(ml_result), risque_label=ml_label, risque_niv=ml_niv,
        ml_confiance=ml_result["confiance_risque"]    if ml_result else None,
        ml_proba_abandon=ml_result["probabilite_abandon"] if ml_result else None,
        ml_niveau_abandon=ml_abandon_label, ml_abandon_css=ml_abandon_css,
        ml_recommandations=ml_result["recommandations"][:3] if ml_result else [],
        ml_source=ml_result["source"] if ml_result else "N/A",
    )


# ==========================
# FILIÈRE
# ==========================
@app.route("/filiere", methods=["GET", "POST"])
@respo_required
def filiere():
    annee_active = get_annee_active()

    filiere_selected  = request.args.get("filiere",  "all")
    semestre_selected = request.args.get("semestre", "all")

    filieres_list = Filiere.query.order_by(Filiere.nom_filiere).all()
    semestres     = get_semestres()

    all_inscriptions_q = _query_inscriptions_avec_relations(
        annee_id=annee_active.id_annee if annee_active else None
    )
    if filiere_selected != "all":
        try:
            all_inscriptions_q = all_inscriptions_q.filter(
                Inscription.id_filiere == int(filiere_selected)
            )
        except ValueError:
            pass
    all_inscriptions = all_inscriptions_q.all()

    all_ml_results = predict_batch(all_inscriptions)
    ml_par_insc    = ml_index(all_inscriptions, all_ml_results)

    insc_par_filiere = defaultdict(list)
    for ins in all_inscriptions:
        insc_par_filiere[ins.id_filiere].append(ins)

    filieres_a_afficher = filieres_list
    if filiere_selected != "all":
        try:
            fid = int(filiere_selected)
            filieres_a_afficher = [f for f in filieres_list if f.id_filiere == fid]
        except ValueError:
            pass

    stats = []
    for f in filieres_a_afficher:
        inscrits_list = insc_par_filiere[f.id_filiere]
        if not inscrits_list:
            continue

        if semestre_selected == "1":
            total   = sum(1 for i in inscrits_list if i.moyenne_s1 is not None)
            reussis = sum(1 for i in inscrits_list if i.moyenne_s1 is not None and i.moyenne_s1 >= 10)
        elif semestre_selected == "2":
            total   = sum(1 for i in inscrits_list if i.moyenne_s2 is not None)
            reussis = sum(1 for i in inscrits_list if i.moyenne_s2 is not None and i.moyenne_s2 >= 10)
        else:
            total   = sum(1 for i in inscrits_list if i.moyenne_annuelle is not None)
            reussis = sum(1 for i in inscrits_list if i.moyenne_annuelle is not None and i.moyenne_annuelle >= 10)

        taux_reussite = round((reussis / total * 100), 2) if total else 0
        if taux_reussite >= 75:   statut = "Excellent"
        elif taux_reussite >= 60: statut = "Actif"
        elif taux_reussite >= 40: statut = "Surveillance"
        else:                     statut = "Critique"

        niveaux_f  = sorted({ins.niveau.libelle for ins in inscrits_list if ins.niveau})
        niveau_str = " - ".join(niveaux_f) if niveaux_f else "-"

        niveaux_ml  = [ml_par_insc[i.id_inscription]["niveau_risque"]      for i in inscrits_list if ml_par_insc.get(i.id_inscription)]
        probas_ml   = [ml_par_insc[i.id_inscription]["probabilite_abandon"] for i in inscrits_list if ml_par_insc.get(i.id_inscription)]
        nb_critiques_f  = sum(1 for n in niveaux_ml if n == "critique")
        nb_moderes_f    = sum(1 for n in niveaux_ml if n == "modere")
        nb_abandons_f   = sum(1 for p in probas_ml  if p >= 0.7)
        proba_abandon_f = round(sum(probas_ml) / len(probas_ml) * 100, 1) if probas_ml else 0

        n_total = len(inscrits_list)
        if nb_critiques_f >= n_total * 0.3:
            niv_ml_filiere = "critique"
        elif (nb_critiques_f + nb_moderes_f) >= n_total * 0.3:
            niv_ml_filiere = "modere"
        else:
            niv_ml_filiere = "ok"

        stats.append({
            "filiere":           f.nom_filiere,
            "niveaux":           niveau_str,
            "inscrits":          len(inscrits_list),
            "taux_reussite":     taux_reussite,
            "statut":            statut,
            "nb_critiques_ml":   nb_critiques_f,
            "nb_moderes_ml":     nb_moderes_f,
            "nb_abandons_ml":    nb_abandons_f,
            "proba_abandon_moy": proba_abandon_f,
            "risque_niv":        niv_ml_filiere,
        })

    return render_template(
        "respo/filiere.html",
        filieres=stats, filieres_list=filieres_list,
        filiere_selected=filiere_selected, semestre=semestres,
        semestre_selected=semestre_selected, annee_active=annee_active,
        title='Filière'
    )


# ==========================
# ALERTES
# ==========================
@app.route("/alerte")
@respo_required
def alerte():
    annee_active      = get_annee_active()
    semestre_selected = request.args.get("semestre", "all")
    niveau_selected   = request.args.get("niveau",   "all")
    filiere_selected  = request.args.get("filiere",  "all")

    query = _query_inscriptions_avec_relations(
        annee_id=annee_active.id_annee if annee_active else None
    )
    if niveau_selected != "all":
        try: query = query.filter(Inscription.id_niveau == int(niveau_selected))
        except ValueError: pass
    if filiere_selected != "all":
        try: query = query.filter(Inscription.id_filiere == int(filiere_selected))
        except ValueError: pass

    if semestre_selected == "1":
        query = query.filter(Inscription.moyenne_s1.isnot(None), Inscription.moyenne_s1 < 10)
    elif semestre_selected == "2":
        query = query.filter(Inscription.moyenne_s2.isnot(None), Inscription.moyenne_s2 < 10)
    else:
        query = query.filter(or_(
            and_(Inscription.moyenne_s1.isnot(None), Inscription.moyenne_s1 < 10),
            and_(Inscription.moyenne_s2.isnot(None), Inscription.moyenne_s2 < 10)
        ))
    inscriptions_a_risque = query.all()

    ml_results_list = predict_batch(inscriptions_a_risque)

    etudiants = []
    for ins, ml_result in zip(inscriptions_a_risque, ml_results_list):
        etu = ins.etudiant
        if semestre_selected == "1":
            moyenne = ins.moyenne_s1
        elif semestre_selected == "2":
            moyenne = ins.moyenne_s2
        else:
            m1 = ins.moyenne_s1 if ins.moyenne_s1 is not None else 20
            m2 = ins.moyenne_s2 if ins.moyenne_s2 is not None else 20
            moyenne = min(m1, m2)

        # ── Crédits validés — FIX parité id_semestre ──────────────────────
        if semestre_selected == "1":
            credits_requis  = 30
            credits_valides = sum(
                r.matiere.credit for r in ins.resultats
                if r.matiere and r.moyenne is not None
                and r.moyenne >= 10 and r.id_semestre % 2 == 1
            )
        elif semestre_selected == "2":
            credits_requis  = 30
            credits_valides = sum(
                r.matiere.credit for r in ins.resultats
                if r.matiere and r.moyenne is not None
                and r.moyenne >= 10 and r.id_semestre % 2 == 0
            )
        else:
            credits_requis  = 60
            credits_valides = sum(
                r.matiere.credit for r in ins.resultats
                if r.matiere and r.moyenne is not None and r.moyenne >= 10
            )

        credits_restants = max(0, credits_requis - credits_valides)
        taux_risque      = round((credits_restants / credits_requis * 100), 2) if credits_requis else 0

        # ── Matières rattrapage — FIX parité id_semestre ──────────────────
        if semestre_selected == "1":
            matieres_rattrapage = sorted({
                res.matiere.nom_matiere for res in ins.resultats
                if res.matiere and res.moyenne is not None
                and res.moyenne < 10 and res.id_semestre % 2 == 1
            })
        elif semestre_selected == "2":
            matieres_rattrapage = sorted({
                res.matiere.nom_matiere for res in ins.resultats
                if res.matiere and res.moyenne is not None
                and res.moyenne < 10 and res.id_semestre % 2 == 0
            })
        else:
            matieres_rattrapage = sorted({
                res.matiere.nom_matiere for res in ins.resultats
                if res.matiere and res.moyenne is not None and res.moyenne < 10
            })

        ml_label, ml_niv                 = ml_badge(ml_result)
        ml_abandon_label, ml_abandon_css = ml_badge_abandon(ml_result)
        ml_score                         = score_ml_to_int(ml_result)

        etudiants.append({
            "matricule":           etu.matricule,
            "nom":                 etu.nom,
            "prenom":              etu.prenom,
            "filiere":             ins.filiere.nom_filiere       if ins.filiere    else "-",
            "specialite":          ins.specialite.nom_specialite if ins.specialite else "-",
            "niveau":              ins.niveau.libelle            if ins.niveau     else "-",
            "moyenne":             moyenne,
            "credits":             credits_valides,
            "credits_requis":      credits_requis,
            "credits_restants":    credits_restants,
            "matieres_rattrapage": matieres_rattrapage,
            "nb_matieres":         len(matieres_rattrapage),
            "taux_risque":         taux_risque,
            "score_risque":        ml_score,
            "niveau_risque":       ml_result["niveau_risque"]        if ml_result else "ok",
            "risque_label":        ml_label,
            "risque_niv":          ml_niv,
            "ml_confiance":        ml_result["confiance_risque"]     if ml_result else None,
            "ml_proba_abandon":    ml_result["probabilite_abandon"]  if ml_result else None,
            "ml_niveau_abandon":   ml_abandon_label,
            "ml_abandon_css":      ml_abandon_css,
            "ml_recommandations":  ml_result["recommandations"][:2]  if ml_result else [],
            "ml_source":           ml_result["source"]               if ml_result else "-",
        })

    etudiants.sort(key=lambda x: (-(x.get("ml_proba_abandon") or 0), -x["score_risque"]))

    critiques_filieres      = sum(1 for e in etudiants if e["niveau_risque"] == "critique")
    moderees_filieres       = sum(1 for e in etudiants if e["niveau_risque"] == "modere")
    nb_abandons_ml          = sum(1 for e in etudiants if (e.get("ml_proba_abandon") or 0) >= 0.7)
    total_alertes           = len(etudiants)
    niveaux_critiques_liste = sorted(set(e["niveau"] for e in etudiants if e["niveau_risque"] == "critique"))
    niveaux_critiques       = len(niveaux_critiques_liste)
    niveaux_moderes_liste   = sorted(set(e["niveau"] for e in etudiants if e["niveau_risque"] == "modere"))
    niveaux_moderes         = len(niveaux_moderes_liste)

    scores_list = [e["score_risque"] for e in etudiants]
    score_moyen = round(sum(scores_list) / len(scores_list), 1) if scores_list else 0
    probas_list = [e["ml_proba_abandon"] for e in etudiants if e["ml_proba_abandon"] is not None]
    proba_moy   = round(sum(probas_list) / len(probas_list) * 100, 1) if probas_list else 0

    if semestre_selected == "1":   condition = Inscription.moyenne_s1 < 10
    elif semestre_selected == "2": condition = Inscription.moyenne_s2 < 10
    else:                          condition = or_(Inscription.moyenne_s1 < 10, Inscription.moyenne_s2 < 10)

    niveau_agg_q = db.session.query(
        Inscription.id_niveau, Niveau.libelle,
        func.count(Inscription.id_inscription).label("total"),
        func.sum(case((condition, 1), else_=0)).label("echecs"),
    ).join(Niveau, Inscription.id_niveau == Niveau.id_niveau)
    if annee_active: niveau_agg_q = niveau_agg_q.filter(Inscription.id_annee == annee_active.id_annee)
    if niveau_selected != "all":
        try: niveau_agg_q = niveau_agg_q.filter(Inscription.id_niveau == int(niveau_selected))
        except ValueError: pass
    if filiere_selected != "all":
        try: niveau_agg_q = niveau_agg_q.filter(Inscription.id_filiere == int(filiere_selected))
        except ValueError: pass
    niveau_agg = niveau_agg_q.group_by(Inscription.id_niveau, Niveau.libelle).all()

    filiere_agg_q = db.session.query(
        Inscription.id_filiere,
        func.count(Inscription.id_inscription).label("total"),
        func.sum(case((condition, 1), else_=0)).label("echecs"),
    )
    if annee_active: filiere_agg_q = filiere_agg_q.filter(Inscription.id_annee == annee_active.id_annee)
    if niveau_selected != "all":
        try: filiere_agg_q = filiere_agg_q.filter(Inscription.id_niveau == int(niveau_selected))
        except ValueError: pass
    if filiere_selected != "all":
        try: filiere_agg_q = filiere_agg_q.filter(Inscription.id_filiere == int(filiere_selected))
        except ValueError: pass
    filiere_agg = filiere_agg_q.group_by(Inscription.id_filiere).all()

    risques_q = (
        db.session.query(Matiere.id_matiere)
        .join(Resultat,    Resultat.id_matiere        == Matiere.id_matiere)
        .join(Inscription, Inscription.id_inscription == Resultat.id_inscription)
        .filter(Resultat.moyenne.isnot(None))
    )
    if annee_active: risques_q = risques_q.filter(Inscription.id_annee == annee_active.id_annee)
    if semestre_selected in ("1", "2"):
        risques_q = risques_q.filter(Resultat.id_semestre % 2 == int(semestre_selected) % 2)
    if niveau_selected != "all":
        try: risques_q = risques_q.filter(Inscription.id_niveau == int(niveau_selected))
        except ValueError: pass
    if filiere_selected != "all":
        try: risques_q = risques_q.filter(Inscription.id_filiere == int(filiere_selected))
        except ValueError: pass
    risques = risques_q.group_by(Matiere.id_matiere)\
        .having(func.sum(case((Resultat.moyenne < 10, 1), else_=0)) > 0).count()

    abandon_q = db.session.query(func.count(Inscription.id_inscription))\
        .filter(Inscription.moyenne_s1.is_(None), Inscription.moyenne_s2.is_(None))
    if annee_active: abandon_q = abandon_q.filter(Inscription.id_annee == annee_active.id_annee)
    if niveau_selected != "all":
        try: abandon_q = abandon_q.filter(Inscription.id_niveau == int(niveau_selected))
        except ValueError: pass
    if filiere_selected != "all":
        try: abandon_q = abandon_q.filter(Inscription.id_filiere == int(filiere_selected))
        except ValueError: pass
    abandon = abandon_q.scalar() or 0

    filiere_objs = {f.id_filiere: f for f in Filiere.query.all()}

    abandon_par_filiere_q = db.session.query(
        Inscription.id_filiere,
        func.count(Inscription.id_inscription).label("nb_abandons"),
    ).filter(Inscription.moyenne_s1.is_(None), Inscription.moyenne_s2.is_(None))
    if annee_active:
        abandon_par_filiere_q = abandon_par_filiere_q.filter(Inscription.id_annee == annee_active.id_annee)
    abandon_par_filiere_agg = {
        row.id_filiere: row.nb_abandons
        for row in abandon_par_filiere_q.group_by(Inscription.id_filiere).all()
    }

    for row in filiere_agg:
        f_obj         = filiere_objs.get(row.id_filiere)
        nom_f         = f_obj.nom_filiere if f_obj else f"Filière {row.id_filiere}"
        nb_abandons_f = abandon_par_filiere_agg.get(row.id_filiere, 0)
        if nb_abandons_f > 0:
            upsert_alerte(f"abandon_filiere_{row.id_filiere}",
                f"Abandon détecté : {nb_abandons_f} étudiant(s) en abandon · {nom_f}")

    if total_alertes > 0:
        msg = (f"{total_alertes} étudiant(s) à risque · {critiques_filieres} critiques · "
               f"{nb_abandons_ml} à fort risque d'abandon ({proba_moy:.0f}%)")
        if annee_active: msg += f" · {annee_active.libelle}"
        upsert_alerte("etudiants_a_risque", msg)

    for row in filiere_agg:
        if (row.total or 0) > 0 and (float(row.echecs or 0) / float(row.total)) >= 0.5:
            f_obj = filiere_objs.get(row.id_filiere)
            nom_f = f_obj.nom_filiere if f_obj else f"Filière {row.id_filiere}"
            taux  = round((float(row.echecs or 0) / float(row.total)) * 100, 1)
            upsert_alerte(f"echec_critique_filiere_{row.id_filiere}",
                f"Taux d'échec critique : {taux}% d'échec dans la filière {nom_f}")

    for row in niveau_agg:
        if (row.total or 0) > 0:
            taux_r = (float(row.total) - float(row.echecs or 0)) / float(row.total)
            if taux_r < 0.3:
                upsert_alerte(f"faible_reussite_niveau_{row.id_niveau}",
                    f"Faible réussite : seulement {round(taux_r * 100, 1)}% d'admis au niveau {row.libelle}")

    matieres_critiques_q = (
        db.session.query(
            Matiere.id_matiere, Matiere.nom_matiere,
            func.count(Resultat.id_resultat).label("total"),
            func.sum(case((Resultat.moyenne < 10, 1), else_=0)).label("echecs")
        )
        .join(Resultat,    Resultat.id_matiere     == Matiere.id_matiere)
        .join(Inscription, Resultat.id_inscription == Inscription.id_inscription)
        .filter(Resultat.moyenne.isnot(None))
    )
    if annee_active: matieres_critiques_q = matieres_critiques_q.filter(Inscription.id_annee == annee_active.id_annee)
    if filiere_selected != "all":
        try: matieres_critiques_q = matieres_critiques_q.filter(Inscription.id_filiere == int(filiere_selected))
        except ValueError: pass
    for row in matieres_critiques_q.group_by(Matiere.id_matiere, Matiere.nom_matiere).all():
        if (row.total or 0) >= 5 and (float(row.echecs or 0) / float(row.total)) >= 0.7:
            taux = round((float(row.echecs or 0) / float(row.total)) * 100, 1)
            upsert_alerte(f"matiere_critique_{row.id_matiere}",
                f"Matière critique : {taux}% d'échec en {row.nom_matiere}")

    for row in filiere_agg:
        if int(row.total or 0) < 5:
            f_obj = filiere_objs.get(row.id_filiere)
            nom_f = f_obj.nom_filiere if f_obj else f"Filière {row.id_filiere}"
            upsert_alerte(f"effectif_faible_{row.id_filiere}",
                f"Effectif faible : seulement {int(row.total or 0)} étudiant(s) inscrit(s) en {nom_f}")

    redoublants_q = db.session.query(func.count(Inscription.id_inscription))\
        .filter(Inscription.est_redoublant == True)
    if annee_active: redoublants_q = redoublants_q.filter(Inscription.id_annee == annee_active.id_annee)
    if filiere_selected != "all":
        try: redoublants_q = redoublants_q.filter(Inscription.id_filiere == int(filiere_selected))
        except ValueError: pass
    nb_redoublants = redoublants_q.scalar() or 0
    if nb_redoublants > 0:
        upsert_alerte("redoublants",
            f"{nb_redoublants} étudiant(s) redoublant(s) détecté(s)"
            + (f" · {annee_active.libelle}" if annee_active else ""))

    if nb_abandons_ml > 0:
        upsert_alerte("ml_abandon_probable",
            f"ML : {nb_abandons_ml} étudiant(s) à fort risque d'abandon (probabilité ≥ 70%)")
    else:
        delete_alerte("ml_abandon_probable")

    db.session.commit()

    nb_alertes    = db.session.query(func.count(Alerte.id_alerte)).scalar() or 0
    semestres     = get_semestres()
    niveaux       = Niveau.query.all()
    filieres_list = Filiere.query.all()

    return render_template(
        "respo/alerte.html",
        etudiants=etudiants, critiques=critiques_filieres, moderees=moderees_filieres,
        total_alertes=total_alertes, niveaux_critiques=niveaux_critiques,
        niveaux_critiques_liste=niveaux_critiques_liste, niveaux_moderes=niveaux_moderes,
        niveaux_moderes_liste=niveaux_moderes_liste, risques=risques, abandon=abandon,
        etudiants_risque=total_alertes, score_moyen=score_moyen,
        nb_abandons_ml=nb_abandons_ml, proba_abandon_moy=proba_moy,
        nb_alertes=nb_alertes, semestre=semestres, semestre_selected=semestre_selected,
        niveau_selected=niveau_selected, filiere_selected=filiere_selected,
        niveaux=niveaux, filieres_list=filieres_list, annee_active=annee_active,
        ml_source=ml_engine.status().get("risk_model_trained") and "ml" or "heuristic",
        title='Alertes & Risques'
    )


# ==========================
# DÉTAIL RISQUE ÉTUDIANT
# ==========================
@app.route("/alerte/detail/<matricule>")
@respo_required
def detail_risque(matricule):
    etudiant = Etudiant.query.filter_by(matricule=matricule).first_or_404()

    inscription = (
        Inscription.query
        .filter_by(id_etudiant=etudiant.id_etudiant)
        .options(
            joinedload(Inscription.niveau).joinedload(Niveau.niveau_suivant),
            joinedload(Inscription.niveau).selectinload(Niveau.semestres),  # ← FIX
            selectinload(Inscription.resultats).joinedload(Resultat.matiere),
        )
        .join(AnneeScolaire, Inscription.id_annee == AnneeScolaire.id_annee)
        .filter(AnneeScolaire.active == True).first()
    )
    if not inscription:
        inscription = (
            Inscription.query
            .filter_by(id_etudiant=etudiant.id_etudiant)
            .options(
                joinedload(Inscription.niveau).joinedload(Niveau.niveau_suivant),
                joinedload(Inscription.niveau).selectinload(Niveau.semestres),  # ← FIX
                selectinload(Inscription.resultats).joinedload(Resultat.matiere),
            )
            .order_by(Inscription.id_inscription.desc()).first()
        )
    if not inscription:
        return "Inscription introuvable", 404

    ml_result                        = predict_ml(inscription)
    ml_label, ml_niv                 = ml_badge(ml_result)
    ml_abandon_label, ml_abandon_css = ml_badge_abandon(ml_result)

    if ml_result:
        niveau_risque_final = ml_result["niveau_risque"]
        proba_abandon       = ml_result["probabilite_abandon"]
        niveau_abandon      = ml_result["niveau_abandon"]
        confiance           = ml_result["confiance_risque"]
        ml_source           = ml_result["source"]
        recommandations_ind = ml_result["recommandations"]
    else:
        sr_fallback         = score_risque(inscription)
        niveau_risque_final = niveau_from_score(sr_fallback)
        ml_label, ml_niv    = badge_risque(sr_fallback)
        proba_abandon       = None
        niveau_abandon      = None
        confiance           = None
        ml_source           = "heuristic"
        recommandations_ind = _build_recos_heuristiques(inscription, ml_niv)

    facteurs = _build_facteurs(inscription)

    matieres_echouees = [
        {"nom": r.matiere.nom_matiere, "moyenne": r.moyenne, "credits": r.matiere.credit}
        for r in inscription.resultats
        if r.matiere and r.moyenne is not None and r.moyenne < 10
    ]
    matieres_echouees.sort(key=lambda x: x["moyenne"])

    credits_req = inscription.niveau.credits_requis if inscription.niveau else 60
    credits_val = sum(
        r.matiere.credit for r in inscription.resultats
        if r.matiere and r.moyenne is not None and r.moyenne >= 10
    )
    annee_active = get_annee_active()

    return render_template(
        "respo/detail_risque.html",
        etudiant=etudiant, inscription=inscription,
        score_risque=score_ml_to_int(ml_result), risque_label=ml_label, risque_niv=ml_niv,
        ml_confiance=confiance, ml_proba_abandon=proba_abandon,
        ml_niveau_abandon=ml_abandon_label, ml_abandon_css=ml_abandon_css,
        ml_source=ml_source, facteurs=facteurs, matieres_echouees=matieres_echouees,
        recommandations_ind=recommandations_ind, credits_val=credits_val,
        credits_req=credits_req, annee_active=annee_active,
        title=f"Détail risque — {etudiant.nom} {etudiant.prenom}"
    )


def _build_facteurs(inscription):
    facteurs = []
    m = inscription.moyenne_annuelle
    if m is None:
        m = inscription.moyenne_s1 or inscription.moyenne_s2

    if m is not None:
        if m < 6:
            facteurs.append({"icone": "bi-exclamation-triangle-fill", "couleur": "#ef4444", "bg": "#fee2e2",
                "titre": "Moyenne très insuffisante", "detail": f"Moyenne de {m}/20 — niveau critique.", "poids": 45})
        elif m < 8:
            facteurs.append({"icone": "bi-exclamation-circle-fill", "couleur": "#ef4444", "bg": "#fee2e2",
                "titre": "Moyenne insuffisante", "detail": f"Moyenne de {m}/20 — en dessous du seuil acceptable.", "poids": 35})
        elif m < 10:
            facteurs.append({"icone": "bi-dash-circle-fill", "couleur": "#f59e0b", "bg": "#fef3c7",
                "titre": "Moyenne en dessous de la moyenne", "detail": f"Moyenne de {m}/20 — proche du seuil d'échec.", "poids": 25})
        elif m < 12:
            facteurs.append({"icone": "bi-info-circle-fill", "couleur": "#3b82f6", "bg": "#dbeafe",
                "titre": "Moyenne légèrement faible", "detail": f"Moyenne de {m}/20 — passable.", "poids": 10})
        else:
            facteurs.append({"icone": "bi-check-circle-fill", "couleur": "#10b981", "bg": "#d1fae5",
                "titre": "Moyenne satisfaisante", "detail": f"Moyenne de {m}/20 — aucun risque.", "poids": 0})
    else:
        facteurs.append({"icone": "bi-question-circle-fill", "couleur": "#f59e0b", "bg": "#fef3c7",
            "titre": "Aucune note enregistrée", "detail": "Aucune moyenne disponible.", "poids": 20})

    credits_req = inscription.niveau.credits_requis if inscription.niveau else 60
    credits_val = sum(r.matiere.credit for r in inscription.resultats
        if r.matiere and r.moyenne is not None and r.moyenne >= 10)
    ratio = credits_val / credits_req if credits_req else 0

    if ratio < 0.25:
        facteurs.append({"icone": "bi-x-circle-fill", "couleur": "#ef4444", "bg": "#fee2e2",
            "titre": "Crédits très insuffisants", "detail": f"{credits_val}/{credits_req} crédits ({round(ratio*100)}%).", "poids": 30})
    elif ratio < 0.50:
        facteurs.append({"icone": "bi-dash-circle-fill", "couleur": "#f59e0b", "bg": "#fef3c7",
            "titre": "Crédits insuffisants", "detail": f"{credits_val}/{credits_req} crédits ({round(ratio*100)}%).", "poids": 20})
    elif ratio < 0.75:
        facteurs.append({"icone": "bi-info-circle-fill", "couleur": "#3b82f6", "bg": "#dbeafe",
            "titre": "Crédits partiellement validés", "detail": f"{credits_val}/{credits_req} crédits ({round(ratio*100)}%).", "poids": 10})
    else:
        facteurs.append({"icone": "bi-check-circle-fill", "couleur": "#10b981", "bg": "#d1fae5",
            "titre": "Crédits bien validés", "detail": f"{credits_val}/{credits_req} crédits ({round(ratio*100)}%).", "poids": 0})

    if getattr(inscription, 'est_redoublant', False):
        facteurs.append({"icone": "bi-arrow-repeat", "couleur": "#ef4444", "bg": "#fee2e2",
            "titre": "Étudiant redoublant", "detail": "Risque accru d'abandon ou d'échec répété.", "poids": 15})
    if inscription.moyenne_s1 is None and inscription.moyenne_s2 is None:
        facteurs.append({"icone": "bi-person-x-fill", "couleur": "#ef4444", "bg": "#fee2e2",
            "titre": "Abandon probable", "detail": "Aucune note ni au S1 ni au S2.", "poids": 10})
    if inscription.moyenne_s1 is not None and inscription.moyenne_s2 is None:
        facteurs.append({"icone": "bi-graph-down", "couleur": "#f59e0b", "bg": "#fef3c7",
            "titre": "Décrochage en cours d'année", "detail": "S1 validé mais aucune note au S2.", "poids": 5})

    return facteurs


def _build_recos_heuristiques(inscription, niv_sr):
    recos = []
    matieres_echouees = [r for r in inscription.resultats if r.matiere and r.moyenne is not None and r.moyenne < 10]

    def add(priorite, titre, action):
        recos.append({"priorite": priorite, "titre": titre, "action": action, "categorie": "Heuristique"})

    if niv_sr == "critique":
        add("critique", "Entretien pédagogique urgent",
            "Convoquer l'étudiant immédiatement pour identifier les causes des difficultés.")
        add("critique", "Plan de remédiation personnalisé",
            "Programme de soutien intensif couvrant toutes les matières échouées.")
    elif niv_sr == "modere":
        add("moderee", "Suivi individuel régulier", "Points mensuels pour suivre la progression.")
        add("moderee", "Soutien ciblé sur les matières échouées",
            f"Concentrer les efforts sur les {len(matieres_echouees)} matière(s) échouée(s).")
    elif niv_sr == "surveillance":
        add("info", "Surveillance préventive", "Informer l'étudiant et l'encourager à solliciter de l'aide.")
    else:
        add("info", "Maintenir les efforts", "L'étudiant est sur la bonne voie. Encourager la régularité.")

    if inscription.moyenne_s1 is not None and inscription.moyenne_s2 is None:
        add("moderee", "Relance urgente pour le S2",
            "Contacter l'étudiant pour comprendre l'absence de notes au S2.")
    return recos


# ==========================
# RECOMMANDATIONS
# ==========================
@app.route("/recommandations")
@respo_required
def recommandations():
    annee_active     = get_annee_active()
    filiere_selected = request.args.get("filiere", "all")
    niveau_selected  = request.args.get("niveau",  "all")

    filieres = Filiere.query.order_by(Filiere.nom_filiere).all()
    niveaux  = Niveau.query.order_by(Niveau.libelle).all()

    q = _query_inscriptions_avec_relations(
        annee_id=annee_active.id_annee if annee_active else None
    )
    if filiere_selected != "all":
        try: q = q.filter(Inscription.id_filiere == int(filiere_selected))
        except ValueError: pass
    if niveau_selected != "all":
        try: q = q.filter(Inscription.id_niveau == int(niveau_selected))
        except ValueError: pass
    inscriptions = q.all()
    total = len(inscriptions)

    ml_results_list = predict_batch(inscriptions)

    reco_counter   = Counter()
    reco_store     = {}
    reco_etudiants = defaultdict(list)

    for ins, ml_r in zip(inscriptions, ml_results_list):
        if not ml_r:
            continue
        for reco in ml_r.get("recommandations", []):
            titre = reco.get("titre", "")
            reco_counter[titre] += 1
            if titre not in reco_store:
                reco_store[titre] = reco
            reco_etudiants[titre].append(ins)

    recommandations_liste = []
    for titre, count in reco_counter.most_common(20):
        reco = reco_store[titre]
        pct  = round(count / total * 100, 1) if total else 0
        recommandations_liste.append({
            "priorite":  reco.get("priorite", "info"),
            "categorie": reco.get("categorie", "Analyse"),
            "titre":     f"{titre} ({pct}% des étudiants concernés)",
            "detail":    f"{count}/{total} étudiant(s) concerné(s) par cette action.",
            "action":    reco.get("action", ""),
        })

    ordre = {"critique": 0, "moderee": 1, "info": 2}
    recommandations_liste.sort(key=lambda x: (
        ordre.get(x["priorite"], 3),
        -reco_counter.get(x["titre"].rsplit(" (", 1)[0], 0)
    ))

    alertes_db = Alerte.query.order_by(Alerte.date.desc()).limit(20).all()
    categories = sorted({r["categorie"] for r in recommandations_liste})
    nb_total   = len(recommandations_liste)

    return render_template(
        "respo/recommandations.html",
        recommandations=recommandations_liste, alertes=alertes_db,
        categories=categories, nb_total=nb_total,
        filieres=filieres, niveaux=niveaux,
        filiere_selected=filiere_selected, niveau_selected=niveau_selected,
        annee_active=annee_active,
        ml_source=ml_engine.status().get("risk_model_trained") and "ml" or "heuristic",
        title="Recommandations"
    )


# ===================================================================
# PARTIE ADMINISTRATEUR SYSTÈME
# ===================================================================

@app.route("/create_admin", methods=["GET", "POST"])
def create_admin():
    if Administrateur_sy.query.count() > 0 and not current_user.is_authenticated:
        flash("Accès non autorisé.", "danger")
        return redirect(url_for("login"))
    form = AdminForm()
    if form.validate_on_submit():
        existing = Administrateur_sy.query.filter_by(email=form.email.data).first()
        if existing:
            flash("Un compte avec cet email existe déjà.", "danger")
            return render_template("create_admin.html", form=form)
        hashed_password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
        admin = Administrateur_sy(
            nom=form.first_name.data, prenom=form.last_name.data,
            email=form.email.data, mot_de_passe=hashed_password,
            genre=form.genre.data, image_file='default.jpg',
        )
        db.session.add(admin); db.session.commit()
        flash("Compte administrateur système créé avec succès.", "success")
        return redirect(url_for("login"))
    return render_template("admin/create_admin.html", form=form, title="Création de compte")


@app.route("/")
def index():
    if Administrateur_sy.query.count() == 0:
        return redirect(url_for('create_admin'))
    return redirect(url_for('login'))


@app.route("/login", methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        if isinstance(current_user, Administrateur_sy):
            return redirect(url_for('dashboard'))
        elif isinstance(current_user, Respo_peda):
            return redirect(url_for('tableau_de_bord'))

    form = Login_adminForm()
    if form.validate_on_submit():
        admin = Administrateur_sy.query.filter_by(email=form.email.data).first()
        if admin and bcrypt.check_password_hash(admin.mot_de_passe, form.password.data):
            login_user(admin)
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        respo = Respo_peda.query.filter_by(email=form.email.data).first()
        if respo and bcrypt.check_password_hash(respo.mot_de_passe, form.password.data):
            login_user(respo)
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('tableau_de_bord'))
        flash("Connexion échouée. Email ou mot de passe incorrect.", "danger")

    image_file = url_for('static', filename='profile_pics/default.jpg')
    return render_template('login.html', title='Connexion', form=form, image_file=image_file)


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


ALLOWED_IMAGE_TYPES = {"JPEG", "PNG", "GIF", "WEBP"}

def save_picture(form_picture):
    random_hex   = secrets.token_hex(8)
    _, f_ext     = os.path.splitext(form_picture.filename)
    picture_fn   = random_hex + f_ext
    picture_path = os.path.join(app.root_path, 'static/profile_pics', picture_fn)
    img = Image.open(form_picture)
    if img.format not in ALLOWED_IMAGE_TYPES:
        raise ValueError(f"Type d'image non autorisé : {img.format}")
    img = img.resize((200, 200), Image.LANCZOS)
    img.save(picture_path)
    return picture_fn


@app.route("/compte", methods=['GET', 'POST'])
@admin_or_respo_required
def compte():
    form = UpdateAccountForm()
    if form.validate_on_submit():
        if form.picture.data:
            try:
                current_user.image_file = save_picture(form.picture.data)
            except ValueError as e:
                flash(str(e), "danger")
                return redirect(url_for('compte'))
        current_user.nom    = form.first_name.data
        current_user.prenom = form.last_name.data
        current_user.email  = form.email.data
        current_user.genre  = form.genre.data
        if form.password.data:
            current_user.mot_de_passe = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
        db.session.commit()
        flash('Votre compte a été mis à jour avec succès', 'success')
        return redirect(url_for('compte'))
    elif request.method == 'GET':
        form.first_name.data = current_user.nom
        form.last_name.data  = current_user.prenom
        form.email.data      = current_user.email
        form.genre.data      = current_user.genre
    image_file = url_for('static', filename='profile_pics/' + current_user.image_file)
    return render_template('compte.html', title='Compte', form=form, image_file=image_file)


@app.route("/parametrage", methods=["GET", "POST"])
@admin_required
def parametrage():
    def parse_date(d):
        try:
            return datetime.strptime(d, "%Y-%m-%d").date() if d else None
        except ValueError:
            return None

    if request.method == "POST":
        action   = request.form.get("action", "creer")
        id_annee = request.form.get("id_annee")
        if action == "modifier":
            annee = db.session.get(AnneeScolaire, id_annee)
            if annee:
                annee.date_debut = parse_date(request.form.get("date_debut")) or annee.date_debut
                annee.date_fin   = parse_date(request.form.get("date_fin"))   or annee.date_fin
                db.session.commit()
                flash(f"Année « {annee.libelle} » mise à jour.", "success")
            return redirect(url_for("parametrage"))

        libelle    = request.form.get("annee_libelle", "").strip()
        date_debut = request.form.get("date_debut")
        date_fin   = request.form.get("date_fin")
        if not libelle or not date_debut or not date_fin:
            flash("Tous les champs sont obligatoires.", "danger")
            return redirect(url_for("parametrage"))
        if AnneeScolaire.query.filter_by(libelle=libelle).first():
            flash(f"L'année « {libelle} » existe déjà.", "danger")
            return redirect(url_for("parametrage"))

        annee_precedente = AnneeScolaire.query.filter_by(active=True).first()
        if annee_precedente and annee_precedente.date_fin >= date.today():
            flash(f"Impossible de créer une nouvelle année : « {annee_precedente.libelle} » est encore active.", "warning")
            return redirect(url_for("parametrage"))

        AnneeScolaire.query.update({"active": False})
        db.session.add(AnneeScolaire(
            libelle=libelle, date_debut=parse_date(date_debut),
            date_fin=parse_date(date_fin), active=True,
        ))

        # ✅ Suppression de toutes les alertes de l'année précédente
        nb_alertes_supprimees = Alerte.query.delete()
        logger.info(
            "Changement d'année scolaire → %d alerte(s) supprimée(s) — nouvelle année : %s",
            nb_alertes_supprimees, libelle
        )

        db.session.commit()

        try:
            ml_engine.auto_train_if_needed(lambda: Inscription.query.all(), force=True)
            flash(f"Année « {libelle} » créée. Modèle ML ré-entraîné.", "success")
        except Exception as e:
            logger.warning("Ré-entraînement ML échoué : %s", e)
            flash(f"Année « {libelle} » créée et activée.", "success")

        return redirect(url_for("parametrage"))

    q     = request.args.get("q", "").strip()
    query = AnneeScolaire.query
    if q:
        date_q = None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try: date_q = datetime.strptime(q, fmt).date(); break
            except ValueError: pass
        if date_q:
            query = query.filter(or_(AnneeScolaire.date_debut == date_q, AnneeScolaire.date_fin == date_q))
        elif q.isdigit():
            query = query.filter(or_(
                extract('year', AnneeScolaire.date_debut) == int(q),
                extract('year', AnneeScolaire.date_fin)   == int(q),
                AnneeScolaire.libelle.ilike(f"%{q}%"),
            ))
        else:
            query = query.filter(or_(
                AnneeScolaire.libelle.ilike(f"%{q}%"),
                cast(AnneeScolaire.date_debut, String).ilike(f"%{q}%"),
                cast(AnneeScolaire.date_fin,   String).ilike(f"%{q}%"),
            ))
    else:
        query = query.order_by(AnneeScolaire.date_debut.desc())

    params           = query.all()
    edit_id          = request.args.get("edit")
    annee_edit       = db.session.get(AnneeScolaire, edit_id) if edit_id else None
    annee_active     = AnneeScolaire.query.filter_by(active=True).first()
    creation_bloquee = annee_active and annee_active.date_fin >= date.today()

    return render_template(
        "admin/parametrage.html",
        params=params, annee_edit=annee_edit, annee_active=annee_active,
        creation_bloquee=creation_bloquee, today=date.today(),
        ml_status=ml_engine.status(), title='Paramétrage académique'
    )

@app.route("/dashboard")
@admin_required
def dashboard():
    users     = Respo_peda.query.all()
    nb_compte = len(users)
    return render_template(
        'admin/dashboard.html',
        nb_compte=nb_compte, users=users,
        ml_status=ml_engine.status(), title='Vue administrative'
    )


@app.route("/ml/train", methods=["POST"])
@admin_required
def ml_train():
    try:
        result = ml_engine.train(Inscription.query.all())
        if result["risk"].get("success"):
            flash(
                f"Modèle ML entraîné — "
                f"{result['meta']['n_samples']} inscriptions · "
                f"Risk F1: {result['risk'].get('cv_f1') or 'N/A'}",
                "success"
            )
        else:
            flash(f"Entraînement échoué — {result['risk'].get('reason', '?')}", "warning")
    except Exception as e:
        flash(f"Erreur lors de l'entraînement ML : {e}", "danger")
    return redirect(url_for("dashboard"))


@app.route("/ml/status")
@admin_or_respo_required
def ml_status():
    from flask import jsonify
    return jsonify(ml_engine.status())


@app.route("/gestion", methods=["GET", "POST"])
@admin_required
def gestion():
    form = RespoForm()
    if form.validate_on_submit():
        existing = Respo_peda.query.filter_by(email=form.email.data).first()
        if existing:
            flash("Un compte avec cet email existe déjà.", "danger")
            return redirect(url_for("gestion"))
        hashed_password = bcrypt.generate_password_hash(form.mot_de_passe.data).decode("utf-8")
        new_user = Respo_peda(
            nom=form.nom.data, prenom=form.prenom.data, email=form.email.data,
            mot_de_passe=hashed_password, role=form.role.data, genre=form.genre.data,
            created_at=datetime.utcnow(),
        )
        db.session.add(new_user)
        db.session.commit()
        flash("Compte responsable pédagogique créé avec succès !", "success")
        return redirect(url_for("gestion"))

    users = Respo_peda.query.all()
    return render_template("admin/gestion.html", form=form, users=users)


@app.route("/delete_user/<int:user_id>", methods=["POST"])
@admin_required
def delete_user(user_id):
    user = Respo_peda.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    flash("Compte supprimé avec succès !", "success")
    return redirect(url_for("gestion"))


# ==========================
# SCHEDULER
# ==========================
def init_scheduler(app):
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        from apscheduler.triggers.cron import CronTrigger
        from IUAInsight.Warehouse.etl_pipeline import ETLPipeline
    except ImportError:
        logger.warning("APScheduler non installé — pip install apscheduler")
        return None

    scheduler = BackgroundScheduler(daemon=True)

    def job_cloture_auto():
        with app.app_context():
            aujourd_hui       = date.today()
            annees_a_cloturer = AnneeScolaire.query.filter(
                AnneeScolaire.active == True,
                AnneeScolaire.date_fin < aujourd_hui
            ).all()
            for annee in annees_a_cloturer:
                logger.info("Clôture automatique : %s", annee.libelle)
                bilan = cloture_annee(annee.id_annee)
                if bilan.get("success"):
                    # ✅ Supprimer toutes les anciennes alertes sauf la clôture
                    Alerte.query.filter(
                        Alerte.type_alerte != f"cloture_auto_{annee.id_annee}"
                    ).delete(synchronize_session=False)

                    upsert_alerte(
                        f"cloture_auto_{annee.id_annee}",
                        f"Clôture automatique {annee.libelle} — "
                        f"{bilan['admis']} admis · {bilan['admis_dettes']} dettes · "
                        f"{bilan['redoublants']} redoublants · {bilan['abandons']} abandons"
                    )
                    db.session.commit()
                    logger.info(
                        "Alertes supprimées à la clôture de l'année %s", annee.libelle
                    )
                else:
                    logger.error("Clôture auto échouée pour %s : %s",
                                 annee.libelle, bilan.get("erreur"))

    def job_sauvegarde_auto():
        with app.app_context():
            faire_sauvegarde_zip()

    def job_etl_nightly():
        with app.app_context():
            annee = get_annee_active()
            success = ETLPipeline().run(annee_id=annee.id_annee if annee else None)
            if success:
                logger.info("ETL nocturne terminé avec succès.")
            else:
                logger.error("ETL nocturne échoué.")

    scheduler.add_job(
        job_cloture_auto,
        trigger=CronTrigger(hour=0, minute=5),
        id="cloture_auto",
        replace_existing=True,
        misfire_grace_time=3600,
    )
    scheduler.add_job(
        job_etl_nightly,
        trigger=CronTrigger(hour=1, minute=0),
        id="etl_nightly",
        replace_existing=True,
        misfire_grace_time=3600,
    )

    scheduler.start()
    logger.info("Scheduler initialisé — ETL 01h00 · clôture 00h05 · sauvegarde dimanche 02h00")
    return scheduler


# ==========================
# IMPORT DE DONNÉES
# ==========================
import pandas as pd
from werkzeug.utils import secure_filename

ALLOWED_EXTENSIONS = {'xlsx', 'csv'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def lire_fichier(file):
    file.seek(0)
    filename = secure_filename(file.filename)
    ext = filename.rsplit('.', 1)[1].lower()
    if ext == 'xlsx':
        return pd.read_excel(file)
    elif ext == 'csv':
        return pd.read_csv(file, sep=None, engine='python')


@app.route('/import', methods=['GET', 'POST'])
@admin_required
def import_donnees():
    if request.method == 'POST':
        type_import = request.form.get('type_import')
        file        = request.files.get('fichier')

        if not file or not allowed_file(file.filename):
            flash("Fichier invalide. Utilisez .xlsx ou .csv", "danger")
            return redirect(url_for('import_donnees'))

        try:
            df = lire_fichier(file)
            df.columns = [c.strip().lower() for c in df.columns]

            erreurs = []
            succes  = 0

            if type_import == 'etudiants':
                colonnes_requises = {'matricule', 'nom', 'prenom'}
                if not colonnes_requises.issubset(set(df.columns)):
                    flash(f"Colonnes manquantes. Requises : {colonnes_requises}", "danger")
                    return redirect(url_for('import_donnees'))

                for i, row in df.iterrows():
                    try:
                        matricule = str(row['matricule']).strip()
                        if not matricule or matricule == 'nan':
                            erreurs.append(f"Ligne {i+2} : matricule vide")
                            continue

                        annee_naissance = None
                        if 'annee_naissance' in df.columns:
                            try:
                                annee_naissance = int(row['annee_naissance'])
                            except (ValueError, TypeError):
                                pass

                        id_nationalite = None
                        if 'id_nationalite' in df.columns:
                            try:
                                id_nationalite = int(row['id_nationalite'])
                            except (ValueError, TypeError):
                                pass

                        if 'pays' in df.columns and id_nationalite is None:
                            pays = str(row.get('pays', '')).strip()
                            if pays and pays != 'nan':
                                nat = Nationalite.query.filter(Nationalite.pays.ilike(pays)).first()
                                if nat:
                                    id_nationalite = nat.id_nationalite

                        existing = Etudiant.query.filter_by(matricule=matricule).first()
                        if existing:
                            existing.nom    = str(row.get('nom', '')).strip()
                            existing.prenom = str(row.get('prenom', '')).strip()
                            if 'genre' in df.columns:
                                existing.genre = str(row.get('genre', '')).strip()
                            if annee_naissance:
                                existing.annee_naissance = annee_naissance
                            if id_nationalite:
                                existing.id_nationalite = id_nationalite
                        else:
                            db.session.add(Etudiant(
                                matricule       = matricule,
                                nom             = str(row.get('nom', '')).strip(),
                                prenom          = str(row.get('prenom', '')).strip(),
                                genre           = str(row.get('genre', '')).strip() if 'genre' in df.columns else None,
                                annee_naissance = annee_naissance,
                                id_nationalite  = id_nationalite,
                            ))
                        succes += 1
                    except Exception as e:
                        erreurs.append(f"Ligne {i+2} : {str(e)}")

            elif type_import == 'inscriptions':
                colonnes_requises = {'matricule', 'id_filiere', 'id_niveau'}
                if not colonnes_requises.issubset(set(df.columns)):
                    flash(f"Colonnes manquantes. Requises : {colonnes_requises}", "danger")
                    return redirect(url_for('import_donnees'))

                annee_active = get_annee_active()
                if not annee_active:
                    flash("Aucune année scolaire active. Créez-en une dans Paramétrage.", "danger")
                    return redirect(url_for('import_donnees'))

                filieres_valides    = {f.id_filiere    for f in Filiere.query.all()}
                niveaux_valides     = {n.id_niveau     for n in Niveau.query.all()}
                specialites_valides = {s.id_specialite for s in Specialite.query.all()}

                for i, row in df.iterrows():
                    try:
                        matricule = str(row['matricule']).strip()
                        if not matricule or matricule == 'nan':
                            erreurs.append(f"Ligne {i+2} : matricule vide")
                            continue

                        etudiant = Etudiant.query.filter_by(matricule=matricule).first()
                        if not etudiant:
                            erreurs.append(f"Ligne {i+2} : étudiant '{matricule}' introuvable")
                            continue

                        try:
                            id_filiere = int(row['id_filiere'])
                        except (ValueError, TypeError):
                            erreurs.append(f"Ligne {i+2} : id_filiere invalide")
                            continue
                        if id_filiere not in filieres_valides:
                            erreurs.append(f"Ligne {i+2} : filière {id_filiere} introuvable")
                            continue

                        try:
                            id_niveau = int(row['id_niveau'])
                        except (ValueError, TypeError):
                            erreurs.append(f"Ligne {i+2} : id_niveau invalide")
                            continue
                        if id_niveau not in niveaux_valides:
                            erreurs.append(f"Ligne {i+2} : niveau {id_niveau} introuvable")
                            continue

                        id_specialite = None
                        if 'id_specialite' in df.columns:
                            try:
                                val = row['id_specialite']
                                if pd.notna(val) and str(val).strip() not in ('', 'nan'):
                                    id_specialite = int(val)
                                    if id_specialite not in specialites_valides:
                                        erreurs.append(f"Ligne {i+2} : spécialité {id_specialite} introuvable")
                                        continue
                            except (ValueError, TypeError):
                                pass

                        est_redoublant = False
                        if 'est_redoublant' in df.columns:
                            try:
                                val = str(row['est_redoublant']).strip().lower()
                                est_redoublant = val in ('1', 'true', 'oui', 'yes', 'o')
                            except Exception:
                                pass

                        existing = Inscription.query.filter_by(
                            id_etudiant = etudiant.id_etudiant,
                            id_annee    = annee_active.id_annee,
                            id_niveau   = id_niveau,
                        ).first()

                        if existing:
                            existing.id_filiere     = id_filiere
                            existing.id_specialite  = id_specialite
                            existing.est_redoublant = est_redoublant
                        else:
                            db.session.add(Inscription(
                                id_etudiant    = etudiant.id_etudiant,
                                id_annee       = annee_active.id_annee,
                                id_filiere     = id_filiere,
                                id_niveau      = id_niveau,
                                id_specialite  = id_specialite,
                                est_redoublant = est_redoublant,
                            ))
                        succes += 1
                    except Exception as e:
                        erreurs.append(f"Ligne {i+2} : {str(e)}")

            elif type_import == 'notes':
                colonnes_requises = {'matricule', 'nom_matiere', 'moyenne'}
                if not colonnes_requises.issubset(set(df.columns)):
                    flash(f"Colonnes manquantes. Requises : {colonnes_requises}", "danger")
                    return redirect(url_for('import_donnees'))

                annee_active = get_annee_active()
                if not annee_active:
                    flash("Aucune année scolaire active.", "danger")
                    return redirect(url_for('import_donnees'))

                matieres_cache = {m.nom_matiere.lower(): m for m in Matiere.query.all()}
                inscriptions_a_recalculer = {}

                for i, row in df.iterrows():
                    try:
                        matricule   = str(row['matricule']).strip()
                        nom_matiere = str(row['nom_matiere']).strip()

                        try:
                            moyenne = float(row['moyenne'])
                        except (ValueError, TypeError):
                            erreurs.append(f"Ligne {i+2} : moyenne invalide")
                            continue

                        if 'id_semestre' not in df.columns:
                            id_semestre = 1
                        else:
                            try:
                                id_semestre = int(row['id_semestre'])
                            except (ValueError, TypeError):
                                erreurs.append(f"Ligne {i+2} : id_semestre invalide")
                                continue

                        semestre = Semestre.query.get(id_semestre)
                        if not semestre:
                            erreurs.append(f"Ligne {i+2} : semestre {id_semestre} introuvable")
                            continue

                        etudiant = Etudiant.query.filter_by(matricule=matricule).first()
                        if not etudiant:
                            erreurs.append(f"Ligne {i+2} : étudiant '{matricule}' introuvable")
                            continue

                        matiere = matieres_cache.get(nom_matiere.lower())
                        if not matiere:
                            erreurs.append(f"Ligne {i+2} : matière '{nom_matiere}' introuvable")
                            continue

                        inscription = Inscription.query.filter_by(
                            id_etudiant = etudiant.id_etudiant,
                            id_annee    = annee_active.id_annee
                        ).first()
                        if not inscription:
                            erreurs.append(f"Ligne {i+2} : aucune inscription pour '{matricule}'")
                            continue

                        if not (0 <= moyenne <= 20):
                            erreurs.append(f"Ligne {i+2} : moyenne {moyenne} hors plage (0-20)")
                            continue

                        existing = Resultat.query.filter_by(
                            id_inscription = inscription.id_inscription,
                            id_matiere     = matiere.id_matiere,
                            id_semestre    = id_semestre
                        ).first()

                        if existing:
                            existing.moyenne       = moyenne
                            existing.credit_valide = moyenne >= 10.0
                        else:
                            db.session.add(Resultat(
                                id_inscription = inscription.id_inscription,
                                id_matiere     = matiere.id_matiere,
                                id_semestre    = id_semestre,
                                moyenne        = moyenne,
                                credit_valide  = moyenne >= 10.0,
                            ))

                        cle = (inscription.id_inscription, id_semestre)
                        inscriptions_a_recalculer[cle] = (inscription, semestre)

                        succes += 1
                    except Exception as e:
                        erreurs.append(f"Ligne {i+2} : {str(e)}")

                db.session.flush()

                for (id_insc, id_sem), (inscription, semestre) in inscriptions_a_recalculer.items():
                    db.session.refresh(inscription)
                    moy = inscription.recalculer_moyenne_semestrielle(id_sem)
                    if semestre.ordre % 2 == 1:
                        inscription.moyenne_s1 = moy
                    else:
                        inscription.moyenne_s2 = moy
                    inscription.recalculer_tout()

            else:
                flash("Type d'import inconnu.", "danger")
                return redirect(url_for('import_donnees'))

            db.session.commit()

            if erreurs:
                flash(f"{succes} importé(s) · {len(erreurs)} erreur(s)", "warning")
                for err in erreurs[:10]:
                    flash(err, "danger")
                if len(erreurs) > 10:
                    flash(f"… et {len(erreurs) - 10} autre(s) erreur(s) non affichée(s).", "warning")
            else:
                flash(f"✅ {succes} ligne(s) importée(s) avec succès !", "success")

        except Exception as e:
            db.session.rollback()
            flash(f"Erreur lecture fichier : {e}", "danger")

        return redirect(url_for('import_donnees'))

    annee_active = get_annee_active()
    filieres     = Filiere.query.order_by(Filiere.nom_filiere).all()
    niveaux      = Niveau.query.order_by(Niveau.libelle).all()
    specialites  = Specialite.query.order_by(Specialite.nom_specialite).all()
    semestres    = Semestre.query.order_by(Semestre.ordre).all()

    return render_template(
        'admin/import.html',
        title        = "Import de données",
        annee_active = annee_active,
        filieres     = filieres,
        niveaux      = niveaux,
        specialites  = specialites,
        semestres    = semestres,
    )


@app.route('/import/template/<type_import>')
@admin_required
def telecharger_template(type_import):
    wb = Workbook()
    ws = wb.active

    H_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    H_FILL  = PatternFill("solid", fgColor="1E2D3D")
    H_ALIGN = Alignment(horizontal="center", vertical="center")
    EX_FONT = Font(name="Arial", size=9, italic=True, color="555555")

    if type_import == 'etudiants':
        ws.title  = "Étudiants"
        headers   = ['matricule', 'nom', 'prenom', 'genre', 'annee_naissance', 'pays']
        exemples  = [
            ['ETU001', 'Kouassi', 'Jean',    'Masculin', 2001, "Côte d'Ivoire"],
            ['ETU002', 'Traoré',  'Aminata', 'Féminin',  2002, 'Mali'],
        ]
        largeurs  = [14, 16, 16, 12, 16, 18]
        notes     = [
            "Obligatoire — identifiant unique", "Obligatoire", "Obligatoire",
            "Masculin / Féminin", "Ex : 2001", "Nom exact du pays (table Nationalite)",
        ]
    elif type_import == 'inscriptions':
        ws.title  = "Inscriptions"
        headers   = ['matricule', 'id_filiere', 'id_niveau', 'id_specialite', 'est_redoublant']
        exemples  = [
            ['ETU001', 1, 2, 1,  'Non'],
            ['ETU002', 1, 2, '', 'Non'],
            ['ETU003', 2, 3, '', 'Oui'],
        ]
        largeurs  = [14, 12, 12, 16, 16]
        notes     = [
            "Doit exister dans la table Etudiant", "ID numérique de la filière",
            "ID numérique du niveau", "Optionnel — laisser vide si aucune spécialité",
            "Oui / Non (défaut : Non)",
        ]
    elif type_import == 'notes':
        ws.title  = "Notes"
        headers   = ['matricule', 'nom_matiere', 'moyenne', 'id_semestre']
        exemples  = [
            ['ETU001', 'Mathématiques', 14.5, 1],
            ['ETU001', 'Informatique',  12.0, 1],
            ['ETU002', 'Mathématiques',  8.5, 2],
        ]
        largeurs  = [14, 24, 10, 14]
        notes     = [
            "Doit exister + être inscrit cette année",
            "Nom exact de la matière (table Matiere)",
            "Entre 0 et 20", "1 = Semestre 1, 2 = Semestre 2",
        ]
    else:
        flash("Type de template inconnu.", "danger")
        return redirect(url_for('import_donnees'))

    ws.row_dimensions[1].height = 22
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = H_FONT
        c.fill      = H_FILL
        c.alignment = H_ALIGN

    for r, ex_row in enumerate(exemples, start=2):
        ws.row_dimensions[r].height = 18
        for col, val in enumerate(ex_row, start=1):
            ws.cell(row=r, column=col, value=val).font = Font(name="Arial", size=9)

    note_row = len(exemples) + 3
    ws.cell(row=note_row, column=1, value="Notes :").font = Font(name="Arial", bold=True, size=9)
    for col, note in enumerate(notes, start=1):
        ws.cell(row=note_row + 1, column=col, value=note).font = EX_FONT

    for col, w in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype      = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment = True,
        download_name = f"template_{type_import}.xlsx"
    )


# ==========================
# ETL MANUEL
# ==========================
@app.route("/etl/run", methods=["POST"])
@admin_required
def etl_run():
    from IUAInsight.Warehouse.etl_pipeline import ETLPipeline
    annee   = get_annee_active()
    success = ETLPipeline().run(annee_id=annee.id_annee if annee else None)
    if success:
        flash("ETL terminé avec succès — Data Warehouse mis à jour.", "success")
    else:
        flash("ETL échoué — vérifiez les logs du serveur.", "danger")
    return redirect(url_for("dashboard"))


@app.route("/etl", methods=["GET", "POST"])
@admin_required
def etl_page():
    from IUAInsight.Warehouse.etl_pipeline import ETLPipeline

    message = None
    succes  = None

    if request.method == "POST":
        action = request.form.get("action")

        if action == "tester_connexion":
            try:
                from sqlalchemy import create_engine, text
                uri    = app.config["SQLALCHEMY_BINDS"]["oltp"]
                engine = create_engine(uri)
                with engine.connect() as conn:
                    conn.execute(text("SELECT 1"))
                message = "✅ Connexion réussie à la base lmd1"
                succes  = True
            except Exception as e:
                message = f"❌ Connexion échouée : {e}"
                succes  = False

        elif action == "lancer_etl":
            annee = get_annee_active()
            try:
                success_etl = ETLPipeline().run(annee_id=annee.id_annee if annee else None)
                if success_etl:
                    message = "✅ ETL terminé avec succès — Data Warehouse mis à jour."
                    succes  = True
                else:
                    message = "❌ ETL échoué — vérifiez les logs."
                    succes  = False
            except Exception as e:
                message = f"❌ Erreur ETL : {e}"
                succes  = False

    annee_active = get_annee_active()
    return render_template(
        "admin/etl.html",
        message=message, succes=succes,
        host="localhost", user="root", db_name="lmd1",
        annee_active=annee_active,
        title="Connexion ETL"
    )

